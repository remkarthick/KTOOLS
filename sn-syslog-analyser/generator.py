when i do run analysis, the below part is visible in browser please fix it and add additional features and additional servicenow error finding rules. table fields shoudl be searchable in header, i shoudl be able to check the errors in multiple instances. there shoudl be a new tab for every instance i configure



.layout{display:flex;height:calc(100vh - 54px);overflow:hidden} .sidebar{width:240px;min-width:200px;background:var(--surface);border-right:1px solid var(--border); display:flex;flex-direction:column;transition:width .25s;overflow:hidden} .sidebar.collapsed{w


#!/usr/bin/env python3
"""
ServiceNow Syslog Analyzer - Project Generator (Flask + SQLite Edition)
Run: python generator.py
"""
import os


def create_file(path: str, content: str):
    dir_ = os.path.dirname(path)
    if dir_:
        os.makedirs(dir_, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  [OK] {path}")


def generate_project():
    print("\n" + "=" * 65)
    print("  ServiceNow Syslog Analyzer — Flask + SQLite Edition")
    print("=" * 65)

    base = "syslog_analyzer"
    for d in [base, f"{base}/templates", f"{base}/static/css",
              f"{base}/static/js", f"{base}/output"]:
        os.makedirs(d, exist_ok=True)
    print(f"\n[+] Folders created under: {base}/\n")

    # ── requirements.txt ──────────────────────────────────────────────────
    create_file(f"{base}/requirements.txt", """\
flask
requests
openpyxl
python-dateutil
""")

    # ── database.py ───────────────────────────────────────────────────────
    create_file(f"{base}/database.py", '''\
"""
database.py — SQLite setup and helpers using Python stdlib sqlite3.
All tables: settings, environments, rules, analysis_runs, analysis_results
"""
import sqlite3
import json
import os
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(__file__), "syslog_analyzer.db")


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    """Create all tables and seed default data if empty."""
    conn = get_conn()
    c = conn.cursor()

    # ── Settings ──────────────────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    """)

    # ── Environments ──────────────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS environments (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            name         TEXT NOT NULL UNIQUE,
            instance_url TEXT NOT NULL,
            username     TEXT NOT NULL,
            password     TEXT NOT NULL,
            table_name   TEXT NOT NULL DEFAULT 'syslog',
            page_size    INTEGER NOT NULL DEFAULT 100,
            timeout      INTEGER NOT NULL DEFAULT 30,
            start_date   TEXT NOT NULL DEFAULT '',
            start_time   TEXT NOT NULL DEFAULT '00:00:00',
            end_date     TEXT NOT NULL DEFAULT '',
            end_time     TEXT NOT NULL DEFAULT '23:59:59',
            level        TEXT NOT NULL DEFAULT '2',
            is_active    INTEGER NOT NULL DEFAULT 0,
            created_at   TEXT NOT NULL,
            updated_at   TEXT NOT NULL
        )
    """)

    # ── Rules ─────────────────────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS rules (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_id            TEXT NOT NULL UNIQUE,
            name               TEXT NOT NULL,
            description        TEXT NOT NULL DEFAULT '',
            pattern            TEXT NOT NULL,
            category           TEXT NOT NULL DEFAULT '',
            severity           TEXT NOT NULL DEFAULT 'Medium',
            source_hint        TEXT NOT NULL DEFAULT '',
            recommendation     TEXT NOT NULL DEFAULT '',
            affected_component TEXT NOT NULL DEFAULT '',
            tags               TEXT NOT NULL DEFAULT '[]',
            is_active          INTEGER NOT NULL DEFAULT 1,
            created_at         TEXT NOT NULL,
            updated_at         TEXT NOT NULL
        )
    """)

    # ── Analysis runs ─────────────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS analysis_runs (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            environment    TEXT NOT NULL,
            start_date     TEXT NOT NULL,
            start_time     TEXT NOT NULL,
            end_date       TEXT NOT NULL,
            end_time       TEXT NOT NULL,
            level          TEXT NOT NULL,
            total_records  INTEGER NOT NULL DEFAULT 0,
            critical_count INTEGER NOT NULL DEFAULT 0,
            high_count     INTEGER NOT NULL DEFAULT 0,
            medium_count   INTEGER NOT NULL DEFAULT 0,
            low_count      INTEGER NOT NULL DEFAULT 0,
            status         TEXT NOT NULL DEFAULT 'pending',
            output_xlsx    TEXT NOT NULL DEFAULT '',
            output_html    TEXT NOT NULL DEFAULT '',
            created_at     TEXT NOT NULL
        )
    """)

    # ── Analysis results ──────────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS analysis_results (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id             INTEGER NOT NULL REFERENCES analysis_runs(id) ON DELETE CASCADE,
            created            TEXT NOT NULL DEFAULT '',
            level              TEXT NOT NULL DEFAULT '',
            message            TEXT NOT NULL DEFAULT '',
            source             TEXT NOT NULL DEFAULT '',
            node               TEXT NOT NULL DEFAULT '',
            job_name           TEXT NOT NULL DEFAULT '',
            user_name          TEXT NOT NULL DEFAULT '',
            thread             TEXT NOT NULL DEFAULT '',
            transaction_id     TEXT NOT NULL DEFAULT '',
            script_record      TEXT NOT NULL DEFAULT '',
            script_table       TEXT NOT NULL DEFAULT '',
            scope              TEXT NOT NULL DEFAULT '',
            page_name          TEXT NOT NULL DEFAULT '',
            record_sys_id      TEXT NOT NULL DEFAULT '',
            root_cause         TEXT NOT NULL DEFAULT '',
            matched_rule_ids   TEXT NOT NULL DEFAULT '',
            matched_rule_names TEXT NOT NULL DEFAULT '',
            categories         TEXT NOT NULL DEFAULT '',
            highest_severity   TEXT NOT NULL DEFAULT '',
            recommendations    TEXT NOT NULL DEFAULT '',
            affected_components TEXT NOT NULL DEFAULT '',
            rules_matched_count INTEGER NOT NULL DEFAULT 0
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_results_run ON analysis_results(run_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_results_sev ON analysis_results(highest_severity)")

    conn.commit()

    # ── Seed default settings ─────────────────────────────────────────────
    defaults = {
        "output_folder":  "output",
        "output_prefix":  "output_error_",
        "output_formats": "xlsx,html",
        "log_level":      "INFO",
        "active_env":     "",
    }
    for k, v in defaults.items():
        c.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)", (k, v))

    # ── Seed default environments ─────────────────────────────────────────
    now = datetime.now().isoformat()
    envs = [
        ("DEV",  "https://your-dev-instance.service-now.com",  "dev_user",  "dev_pass",  0),
        ("UAT",  "https://orangeservicemanagementuat009.service-now.com", "uat_user", "uat_pass", 1),
        ("PROD", "https://your-prod-instance.service-now.com", "prod_user", "prod_pass", 0),
    ]
    for name, url, user, pwd, active in envs:
        c.execute("""
            INSERT OR IGNORE INTO environments
            (name,instance_url,username,password,table_name,page_size,timeout,
             start_date,start_time,end_date,end_time,level,is_active,created_at,updated_at)
            VALUES(?,?,?,?,\'syslog\',100,30,\'2026-07-21\',\'10:34:00\',\'2026-07-21\',\'10:35:00\',\'2\',?,?,?)
        """, (name, url, user, pwd, active, now, now))

    conn.commit()

    # ── Seed default rules ────────────────────────────────────────────────
    _seed_rules(c, now)
    conn.commit()
    conn.close()
    print(f"[DB] Initialised: {DB_PATH}")


def _seed_rules(c, now):
    """Insert 105 default rules if rules table is empty."""
    c.execute("SELECT COUNT(*) as cnt FROM rules")
    if c.fetchone()["cnt"] > 0:
        return

    rules = [
        ("RULE_001","Null Data Property Access","Script attempts to read data property from null",r"Cannot read property.*data.*from null","NullPointerError","High","ResponseHandler script","Add null check before accessing data property. Use: if(response && response.data)","Script Include",'["null","data","responsehandler"]'),
        ("RULE_002","Null Property Read Generic","Generic null property read error",r"Cannot read property.*from null","NullPointerError","High","Any Script","Validate object is not null before property access","Script",'["null","property"]'),
        ("RULE_003","Java NullPointerException","Java-level NullPointerException",r"java\.lang\.NullPointerException","JavaException","Critical","Java layer","Check server-side Java code or raise with ServiceNow support.","Platform",'["java","null","exception"]'),
        ("RULE_004","Invalid GraphQL Query","GraphQL API received null query",r"Invalid query string received:null","APIError","High","api/now/graphql","Ensure GraphQL query is not null before sending.","GraphQL API",'["graphql","null","api"]'),
        ("RULE_005","Role ID Not Found","Flow Designer references non-existent Role ID","Role ID doesn't exist","ConfigurationError","High","Flow Designer","Verify role sys_id in flow exists. Re-map in Flow Designer.","Flow Designer",'["role","flow","configuration"]'),
        ("RULE_006","Processing Framework Job Context Failure","Scheduler failed to retrieve job context",r"ProcessingFramework > Failed to get the schedule job context","SchedulerError","High","ProcessingFramework","Check scheduled job record exists and is active.","Scheduler",'["scheduler","job","context"]'),
        ("RULE_007","Processing Framework Event Processing Failure","Scheduler failed to process an event",r"ProcessingFramework > Failed to process event","SchedulerError","High","ProcessingFramework","Inspect event queue and associated flow for errors.","Scheduler / Event Queue",'["scheduler","event","processing"]'),
        ("RULE_008","Generic Null Error","Generic null error logged",r"\*\*\* ERROR \*\*\* null","GenericError","Medium","System","Investigate scheduled job or script for unhandled null values.","Scheduler / Script",'["null","generic","error"]'),
        ("RULE_009","Invalid Country Code in sys_user","Invalid country code on user record update",r"Attempt to insert/update an invalid value for the country field in sys_user","DataValidationError","Medium","sys_user table","Use ISO 3166-1 alpha-2 codes (e.g. IN not IND).","User Management Script",'["user","country","validation"]'),
        ("RULE_010","Flow Designer Event Processing Failure","Flow Designer failed to process system event",r"Flow Designer.*Failed to process event","FlowDesignerError","High","Flow Designer","Check flow context record and associated flow.","Flow Designer",'["flow","event","failure"]'),
        ("RULE_011","PFSessionClone Role Validation Failure","Flow session clone failed during role validation",r"PFSessionClone.*getValidRoles","FlowDesignerError","High","Flow Designer Engine","Verify roles assigned to flow run-as user are valid.","Flow Designer Engine",'["flow","role","session"]'),
        ("RULE_012","ProcessAutomationException","Process automation exception during flow execution",r"ProcessAutomationException","FlowDesignerError","High","Flow Designer Engine","Review flow definition and run-as user configuration.","Flow Designer Engine",'["flow","automation","exception"]'),
        ("RULE_013","Ebus Stream Reception Error","RecieveEbus stream fromToscan job failing",r"job_name.*RecieveEbus stream fromToscan","IntegrationError","High","Ebus Integration","Check Toscan endpoint availability and Ebus configuration.","Ebus Integration",'["ebus","integration","stream"]'),
        ("RULE_014","Events Process Job Failure","events process 0 job failing to get context",r"job_name.*events process 0","SchedulerError","High","Event Processing Job","Check event processing job configuration and node availability.","Event Processor",'["events","scheduler","job"]'),
        ("RULE_015","Flow Engine Event Handler Failure","Flow Engine Event Handler job error",r"job_name.*Flow Engine Event Handler","FlowDesignerError","High","Flow Engine","Inspect Flow Engine Event Handler job and associated flows.","Flow Engine",'["flow","engine","handler"]'),
        ("RULE_016","Servlet Error Listener Triggered","Servlet error listener caught unhandled exception",r"com\.glide\.ui\.ServletErrorListener","PlatformError","Medium","Servlet Layer","Review full stack trace for root cause.","Servlet",'["servlet","platform","exception"]'),
        ("RULE_017","Scheduler Worker Thread Error","Scheduler worker thread error during job execution",r"com\.glide\.schedule_v2\.SchedulerWorkerThread","SchedulerError","Medium","Scheduler","Check job being executed by worker thread.","Scheduler Worker",'["scheduler","worker","thread"]'),
        ("RULE_018","Script Include Error","Error within a Script Include",r"_script_table.*sys_script_include","ScriptError","High","Script Include","Review Script Include for null handling.","Script Include",'["script","include","error"]'),
        ("RULE_019","System Script Error","Error within a sys_script Business Rule",r"_script_table.*sys_script[^_]","ScriptError","High","Business Rule","Review Business Rule for data validation issues.","Business Rule",'["script","business_rule","error"]'),
        ("RULE_020","Impersonation Active During Error","Error occurred while user was being impersonated",r"_is_impersonating.*true","SecurityWarning","Low","User Session","Verify impersonation was intentional.","User Session",'["impersonation","security","session"]'),
        ("RULE_021","System User Executing Script","Script executed as system user",r"_logged_in_user.*\bsystem\b","InfoPattern","Low","System User","Ensure system-user scripts have proper error handling.","Script",'["system","user","script"]'),
        ("RULE_022","GraphQL Endpoint Error","Error on GraphQL API endpoint",r"_page_name.*api/now/graphql","APIError","High","GraphQL Endpoint","Validate client-side GraphQL query construction.","GraphQL API",'["graphql","api","endpoint"]'),
        ("RULE_023","xmlhttp.do Endpoint Error","Error on xmlhttp.do endpoint",r"_page_name.*xmlhttp\.do","APIError","Medium","xmlhttp.do","Review client-side script making xmlhttp.do call.","AJAX Handler",'["xmlhttp","ajax","endpoint"]'),
        ("RULE_024","Multi-Node Error Propagation","Same error across multiple nodes",r"orangeservicemanagementuat","InfrastructureError","Critical","Multi-Node","Investigate shared configuration affecting all nodes.","Infrastructure",'["multi-node","infrastructure","propagation"]'),
        ("RULE_025","UAT001 Node Error","Error on UAT001 node",r"orangeservicemanagementuat001","NodeError","Medium","UAT001","Check UAT001 node health.","Node UAT001",'["uat001","node"]'),
        ("RULE_026","UAT003 Node Error","Error on UAT003 node",r"orangeservicemanagementuat003","NodeError","Medium","UAT003","Check UAT003 node health.","Node UAT003",'["uat003","node"]'),
        ("RULE_027","UAT006 Node Error","Error on UAT006 node",r"orangeservicemanagementuat006","NodeError","Medium","UAT006","Check UAT006 node health.","Node UAT006",'["uat006","node"]'),
        ("RULE_028","UAT007 Node Error","Error on UAT007 node",r"orangeservicemanagementuat007","NodeError","Medium","UAT007","Check UAT007 node health.","Node UAT007",'["uat007","node"]'),
        ("RULE_029","UAT008 Node Error","Error on UAT008 node",r"orangeservicemanagementuat008","NodeError","Medium","UAT008","Check UAT008 node health.","Node UAT008",'["uat008","node"]'),
        ("RULE_030","UAT009 Node Error","Error on UAT009 node",r"orangeservicemanagementuat009","NodeError","Medium","UAT009","Check UAT009 node health.","Node UAT009",'["uat009","node"]'),
        ("RULE_031","UAT010 Node Error","Error on UAT010 node",r"orangeservicemanagementuat010","NodeError","Medium","UAT010","Check UAT010 node health.","Node UAT010",'["uat010","node"]'),
        ("RULE_032","UAT011 Node Error","Error on UAT011 node",r"orangeservicemanagementuat011","NodeError","Medium","UAT011","Check UAT011 node health.","Node UAT011",'["uat011","node"]'),
        ("RULE_033","P1 ISOS Scheduled Job Error","P1 ISOS ebonded incidents job error",r"Scheduled execution of P1 ISOS New ebonded incidents list","SchedulerError","High","P1 ISOS Job","Review P1 ISOS ebonded incidents script for null handling.","Scheduled Job",'["p1","isos","ebonded","scheduler"]'),
        ("RULE_034","KT Corporation Job Error","KT CORPORATION incident list job error",r"Scheduled execution of P1 KT CORPORATION","SchedulerError","High","KT Corporation Job","Review KT Corporation incident list script.","Scheduled Job",'["kt","corporation","scheduler"]'),
        ("RULE_035","No Thrown Error Suffix","Silent failure no thrown error",r"no thrown error","SilentFailure","Medium","Script","Add explicit try-catch blocks.","Script",'["silent","failure","no_error"]'),
        ("RULE_036","Response Handler Failure","ResponseHandler function failed",r"Error in ResponseHandler","IntegrationError","High","ResponseHandler","Review ResponseHandler script include for null checking.","ResponseHandler Script Include",'["response","handler","integration"]'),
        ("RULE_037","Worker Thread 1 Error","Error on scheduler worker thread 1",r"glide\.scheduler\.worker\.1","SchedulerError","Low","Worker Thread 1","Monitor worker thread 1 for recurring errors.","Scheduler Worker Thread",'["worker","thread","scheduler"]'),
        ("RULE_038","Worker Thread 2 Error","Error on scheduler worker thread 2",r"glide\.scheduler\.worker\.2","SchedulerError","Low","Worker Thread 2","Monitor worker thread 2 for recurring errors.","Scheduler Worker Thread",'["worker","thread","scheduler"]'),
        ("RULE_039","Worker Thread 3 Error","Error on scheduler worker thread 3",r"glide\.scheduler\.worker\.3","SchedulerError","Low","Worker Thread 3","Monitor worker thread 3 for recurring errors.","Scheduler Worker Thread",'["worker","thread","scheduler"]'),
        ("RULE_040","Worker Thread 4 Error","Error on scheduler worker thread 4",r"glide\.scheduler\.worker\.4","SchedulerError","Low","Worker Thread 4","Monitor worker thread 4 for recurring errors.","Scheduler Worker Thread",'["worker","thread","scheduler"]'),
        ("RULE_041","Worker Thread 5 Error","Error on scheduler worker thread 5",r"glide\.scheduler\.worker\.5","SchedulerError","Low","Worker Thread 5","Monitor worker thread 5 for recurring errors.","Scheduler Worker Thread",'["worker","thread","scheduler"]'),
        ("RULE_042","Worker Thread 7 Error","Error on scheduler worker thread 7",r"glide\.scheduler\.worker\.7","SchedulerError","Low","Worker Thread 7","Monitor worker thread 7 for recurring errors.","Scheduler Worker Thread",'["worker","thread","scheduler"]'),
        ("RULE_043","Default HTTP Thread Error","Error on Default HTTP thread",r"Default-thread-","PlatformError","Medium","HTTP Thread","Review HTTP request for invalid input.","HTTP Thread",'["http","thread","default"]'),
        ("RULE_044","Script Record df3c2ee4 Error","Error linked to ResponseHandler script record",r"_script_record.*df3c2ee42b457610c720fe61de91bf98","ScriptError","High","Script Record df3c2ee4","Review Script Include df3c2ee42b457610c720fe61de91bf98.","Script Include",'["script","record","sysid"]'),
        ("RULE_045","Trigger Record Error","Error associated with sys_trigger record",r"table_name.*sys_trigger","SchedulerError","Medium","sys_trigger","Review sys_trigger record configuration.","Trigger",'["trigger","scheduler"]'),
        ("RULE_046","Java Stack Trace Detected","Java stack trace in log entry",r"at com\..*\..*\(.*\.java:\d+\)","JavaException","High","Java Stack","Review recursive calls in identified Java classes.","Java Layer",'["stack","java","trace"]'),
        ("RULE_047","NowMQ Event Handler Error","NowMQ event handler failed",r"NowMQEventHandler","MessagingError","High","NowMQ","Check NowMQ configuration and queue health.","NowMQ",'["nowmq","messaging","event"]'),
        ("RULE_048","NowMQ Processing Framework Job Error","NowMQ processing framework job error",r"NowMQProcessingFrameworkJob","MessagingError","High","NowMQ Processing","Inspect NowMQ processing framework job configuration.","NowMQ Processing",'["nowmq","processing","job"]'),
        ("RULE_049","Flow NowMQ Message Handler Error","Flow Designer NowMQ message handler failed",r"FlowNowMQMessageHandler","FlowDesignerError","High","Flow NowMQ","Review Flow Designer NowMQ message handler configuration.","Flow Designer NowMQ",'["flow","nowmq","handler"]'),
        ("RULE_050","Process Hub Event Handler Error","Process Hub event handler failed",r"ProcessHubEventHandler","FlowDesignerError","High","Process Hub","Review Process Hub event handler and flow configuration.","Process Hub",'["process_hub","event","handler"]'),
        ("RULE_051","Job Executor Error","Job executor error during execution",r"com\.glide\.schedule\.JobExecutor","SchedulerError","High","Job Executor","Review job being executed and its scripts.","Job Executor",'["job","executor","scheduler"]'),
        ("RULE_052","Data Replication Context Error","Error in data replication context",r"DataReplicationAdvisors","ReplicationError","High","Data Replication","Check data replication configuration.","Data Replication",'["replication","data","context"]'),
        ("RULE_053","Transaction Processor Error","Transaction processor error",r"TransactionProcessor","PlatformError","Medium","Transaction Processor","Review transaction being processed.","Transaction Processor",'["transaction","processor","platform"]'),
        ("RULE_054","Glide Process Automation Error","GlideProcessAutomation error",r"GlideProcessAutomation","FlowDesignerError","High","Glide Process Automation","Review flow and run-as user configuration.","Glide Process Automation",'["glide","process","automation"]'),
        ("RULE_055","Sync Execution Lock Error","SyncExecution failed to acquire lock",r"SyncExecution\.tryLock","ConcurrencyError","High","Sync Execution","Check for deadlocks or long-running transactions.","Sync Execution",'["sync","lock","concurrency"]'),
        ("RULE_056","Glide PF Session Error","GlidePFSession error during flow session init",r"GlidePFSession","FlowDesignerError","High","PF Session","Review flow session configuration and run-as user.","PF Session",'["pf","session","flow"]'),
        ("RULE_057","Repeated Error Same Transaction","Multiple errors under same transaction ID",r"_txid","TransactionError","Medium","Transaction","Investigate full transaction for cascading errors.","Transaction",'["transaction","repeated","error"]'),
        ("RULE_058","Script Source Error","Error sourced from Script prefix",r"\*\*\* Script","ScriptError","High","Script","Review script referenced in log entry.","Script",'["script","source","error"]'),
        ("RULE_059","Oceane System User Error","Error under oceane.system user context",r"_logged_in_user.*oceane\.system","UserContextError","Medium","oceane.system user","Review scripts running under oceane.system.","User Context",'["oceane","system","user"]'),
        ("RULE_060","External Orange User Error","Error triggered by orange.com user",r"_logged_in_user.*@orange\.com","UserError","Medium","External User","Notify user and review action that triggered error.","User Interface",'["user","external","orange"]'),
        ("RULE_061","Jasmeen Shaik User Error","Error triggered by jasmeen.shaik",r"jasmeen\.shaik","UserError","Medium","jasmeen.shaik","Review action performed by jasmeen.shaik.","User Interface",'["user","jasmeen","shaik"]'),
        ("RULE_062","Srushtisunil Ingalkar User Error","Error triggered by srushtisunil.ingalkar",r"srushtisunil\.ingalkar","UserError","Medium","srushtisunil.ingalkar","Review GraphQL query construction in client.","User Interface",'["user","srushtisunil","ingalkar"]'),
        ("RULE_063","Karthick Krishnan User Error","Error associated with karthick.krishnan",r"karthick\.krishnan","UserError","Medium","karthick.krishnan","Review flows created by karthick.krishnan for role issues.","Flow Designer",'["user","karthick","krishnan"]'),
        ("RULE_064","Aaishah Bassam Data Error","Invalid data update for aaishah.bassam",r"aaishah\.bassam","DataValidationError","Medium","aaishah.bassam","Correct country code for aaishah.bassam to valid ISO code.","User Management",'["user","aaishah","data"]'),
        ("RULE_065","Global Scope Script Error","Error in global scope script",r"_scope.*global","ScriptError","Medium","Global Scope","Review global scope scripts for null checks.","Global Scope",'["global","scope","script"]'),
        ("RULE_066","Flow Context Record Error","Error referencing flow context record",r"sys_flow_context","FlowDesignerError","High","Flow Context","Review flow context record and retry flow.","Flow Context",'["flow","context","record"]'),
        ("RULE_067","Sysevent Reference Error","Error references sysevent record",r"sysevent\[","EventError","High","System Event","Review sysevent record and processing flow.","System Event",'["sysevent","event","flow"]'),
        ("RULE_068","Repeated Ebus Error Pattern","RecieveEbus stream fromToscan error repeating",r"RecieveEbus stream fromToscan","RecurringError","Critical","Recurring Pattern","Implement circuit breaker. Investigate Ebus stream failure.","Ebus Integration",'["recurring","ebus","pattern"]'),
        ("RULE_069","IND Country Code Invalid","Country code IND not valid ISO alpha-2",r"invalid value for the country field.*IND","DataValidationError","Medium","Country Field","Replace IND with IN in user update script.","User Management",'["country","IND","validation"]'),
        ("RULE_070","Flow Engine Queue Error","Error in flow_engine queue",r"queue=flow_engine","FlowDesignerError","High","Flow Engine Queue","Check flow_engine queue for backlog.","Flow Engine Queue",'["flow","engine","queue"]'),
        ("RULE_071","Scheduler Job Record Missing","Scheduler cannot find job record",r"Failed to get the schedule job context","SchedulerError","High","Scheduler","Verify scheduled job record exists and is active.","Scheduler",'["scheduler","job","missing"]'),
        ("RULE_072","Role Misconfiguration in Flow","Flow configured with non-existent role","Role ID doesn't exist.*b0feb4bf5b1110100206cb837b81c788","ConfigurationError","Critical","Flow Role Configuration","Update flow to use valid role. Role b0feb4bf does not exist.","Flow Designer",'["role","flow","misconfiguration"]'),
        ("RULE_073","Ebus Stream Null Property Failure","Ebus job failing with null property error",r"RecieveEbus.*Cannot read property","IntegrationError","Critical","Ebus Integration","Investigate Toscan endpoint availability immediately.","Ebus / Toscan Integration",'["ebus","toscan","critical"]'),
        ("RULE_074","Session ID Reuse Across Errors","Same session ID in multiple error entries",r"_session_id.*36B6C1542BDE83503A14F60D4391BFF9","SessionError","Medium","User Session","Review user session for repeated failed requests.","Session Management",'["session","reuse","error"]'),
        ("RULE_075","Glide Processing Framework Error","Glide processing framework error",r"com\.glide\.processing\.framework","PlatformError","High","Processing Framework","Review processing framework configuration.","Processing Framework",'["processing","framework","glide"]'),
        ("RULE_076","Script Include Null Return","Script Include returned null",r"sys_script_include.*null","ScriptError","High","Script Include","Ensure Script Include methods return valid objects.","Script Include",'["script","include","null"]'),
        ("RULE_077","Multiple Nodes Same Job Failure","Same job failing on multiple nodes",r"job=aaa5a52f2b3c0b10dc25f2c0ce91bf81","InfrastructureError","Critical","Multi-Node Job Failure","Investigate shared job record aaa5a52f for config issues.","Scheduler / Infrastructure",'["multi-node","job","critical"]'),
        ("RULE_078","Trigger Record Shared Across Nodes","Same trigger record causing errors on multiple nodes",r"record_sys_id.*dd1c35c12bd0c31045fff0845e91bffa","ConfigurationError","High","Trigger Record","Review trigger record dd1c35c1 for configuration issues.","Trigger",'["trigger","shared","nodes"]'),
        ("RULE_079","Glide Schedule V2 Error","Error in Glide Schedule V2 framework",r"com\.glide\.schedule_v2","SchedulerError","Medium","Glide Schedule V2","Review Glide Schedule V2 configuration.","Glide Schedule V2",'["schedule","v2","glide"]'),
        ("RULE_080","Transactional Worker Thread Error","Transactional worker thread error",r"TransactionalWorkerThread","PlatformError","Medium","Transactional Worker","Review transaction being processed by worker thread.","Transactional Worker",'["transactional","worker","thread"]'),
        ("RULE_081","API Endpoint Null Query","API endpoint received null query parameter",r"received:null","APIError","High","API Endpoint","Validate all query parameters before sending to API.","API",'["api","null","query"]'),
        ("RULE_082","Flow Designer Java Stack Trace","Flow Designer error with Java stack trace",r"Flow Designer.*at com\.snc\.process_flow","FlowDesignerError","Critical","Flow Designer","Analyze full stack trace for root cause in Flow Designer engine.","Flow Designer Engine",'["flow","stack","trace"]'),
        ("RULE_083","Glide Worker Error","Glide worker error",r"com\.glide\.worker","PlatformError","Medium","Glide Worker","Review Glide worker configuration.","Glide Worker",'["glide","worker","platform"]'),
        ("RULE_084","Ebonded Incident Job Null Error","Ebonded incident job returned null",r"ebonded incidents.*null|null.*ebonded","IntegrationError","High","Ebonded Integration","Review ebonded incident integration script.","Ebonded Integration",'["ebonded","incident","null"]'),
        ("RULE_085","Job Executor Lambda Error","Lambda in JobExecutor error",r"JobExecutor\.lambda","JavaException","High","Job Executor Lambda","Review lambda expression in JobExecutor.","Job Executor",'["lambda","executor","java"]'),
        ("RULE_086","Message Flow Error","messageFlow method in GlideProcessAutomation failed",r"GlideProcessAutomation\.messageFlow","FlowDesignerError","High","Message Flow","Review message flow configuration in affected flow.","Flow Designer",'["message","flow","automation"]'),
        ("RULE_087","Message Flow Sync Error","messageFlowSync method failed",r"GlideProcessAutomation\.messageFlowSync","FlowDesignerError","High","Message Flow Sync","Review synchronous message flow configuration.","Flow Designer",'["message","flow","sync"]'),
        ("RULE_088","Process Hub Send Message Error","doSendMessage in ProcessHubEventHandler failed",r"ProcessHubEventHandler\.doSendMessage","FlowDesignerError","High","Process Hub Send Message","Review message sending configuration in Process Hub.","Process Hub",'["process_hub","send","message"]'),
        ("RULE_089","Flow NowMQ Process Event Error","processEvent in FlowNowMQMessageHandler failed",r"FlowNowMQMessageHandler\.processEvent","MessagingError","High","Flow NowMQ Process Event","Review NowMQ event processing configuration.","Flow NowMQ",'["nowmq","process","event"]'),
        ("RULE_090","PF Session Clone Initialize Error","initializeRunAsSession in PFSessionClone failed",r"PFSessionClone\.initializeRunAsSession","FlowDesignerError","High","PF Session Clone","Verify run-as user and role configuration for flow.","PF Session Clone",'["pf","session","initialize"]'),
        ("RULE_091","Processing Framework Entities Error","processEntities in ProcessingFrameworkJob failed",r"ProcessingFrameworkJob\.processEntities","PlatformError","High","Processing Framework Entities","Review processing framework job entity configuration.","Processing Framework",'["processing","entities","framework"]'),
        ("RULE_092","Scheduler Execute Job Error","executeJob in SchedulerWorkerThread failed",r"SchedulerWorkerThread\.executeJob","SchedulerError","High","Scheduler Execute Job","Review job being executed by scheduler worker thread.","Scheduler",'["scheduler","execute","job"]'),
        ("RULE_093","Scheduler Process Error","process method in SchedulerWorkerThread failed",r"SchedulerWorkerThread\.process","SchedulerError","High","Scheduler Process","Review scheduler worker thread process method.","Scheduler",'["scheduler","process","worker"]'),
        ("RULE_094","Scheduler Run Error","run method in SchedulerWorkerThread failed",r"SchedulerWorkerThread\.run","SchedulerError","High","Scheduler Run","Review scheduler worker thread run method.","Scheduler",'["scheduler","run","worker"]'),
        ("RULE_095","Job Executor Execute Error","execute method in JobExecutor failed",r"JobExecutor\.execute","SchedulerError","High","Job Executor Execute","Review job executor execute method.","Job Executor",'["executor","execute","scheduler"]'),
        ("RULE_096","Data Replication Context Lambda Error","inDataReplicationContext lambda failed",r"JobExecutor\.lambda\$inDataReplicationContext","ReplicationError","High","Data Replication Context","Review data replication context in job executor.","Job Executor / Replication",'["replication","context","lambda"]'),
        ("RULE_097","Execute Job Lambda Error","executeJob lambda in JobExecutor failed",r"JobExecutor\.lambda\$executeJob","SchedulerError","High","Execute Job Lambda","Review executeJob lambda in JobExecutor.","Job Executor",'["lambda","execute","job"]'),
        ("RULE_098","Process Lambda Error","process lambda in SchedulerWorkerThread failed",r"SchedulerWorkerThread\.lambda\$process","SchedulerError","High","Process Lambda","Review process lambda in SchedulerWorkerThread.","Scheduler",'["lambda","process","scheduler"]'),
        ("RULE_099","Message Flow Lambda Error","messageFlow lambda in GlideProcessAutomation failed",r"GlideProcessAutomation\.lambda\$messageFlow","FlowDesignerError","High","Message Flow Lambda","Review messageFlow lambda in GlideProcessAutomation.","Flow Designer",'["lambda","message","flow"]'),
        ("RULE_100","Process Automation Wrapper Error","GlideProcessAutomationWrapper error",r"GlideProcessAutomationWrapper","FlowDesignerError","High","Process Automation Wrapper","Review process automation wrapper configuration.","Process Automation Wrapper",'["wrapper","automation","flow"]'),
        ("RULE_101","Process Hub Process Event Error","processEvent in ProcessHubEventHandler failed",r"ProcessHubEventHandler\.processEvent","FlowDesignerError","High","Process Hub Process Event","Review processEvent method in ProcessHubEventHandler.","Process Hub",'["process_hub","event","process"]'),
        ("RULE_102","Null Response from External Service","External service returned null response",r"Cannot read property.*from null","IntegrationError","High","External Service","Add null response handling. Check service availability.","Integration",'["null","response","external"]'),
        ("RULE_103","Error Without Stack Trace","Error logged without stack trace",r"\*\*\* ERROR \*\*\* null","DiagnosticIssue","Medium","System","Enable verbose logging to capture full stack traces.","Logging",'["error","no_stack","diagnostic"]'),
        ("RULE_104","Glide Schedule Job Executor Error","com.glide.schedule.JobExecutor error",r"com\.glide\.schedule\.JobExecutor","SchedulerError","High","Job Executor","Review job executor and job being executed.","Job Executor",'["job","executor","glide"]'),
        ("RULE_105","Country Field Validation Error","Field validation error for country field",r"invalid value for the country field","DataValidationError","Medium","Field Validation","Ensure country field values conform to ISO 3166-1 alpha-2 standard.","Field Validation",'["country","field","validation"]'),
    ]

    for r in rules:
        c.execute("""
            INSERT OR IGNORE INTO rules
            (rule_id,name,description,pattern,category,severity,source_hint,
             recommendation,affected_component,tags,is_active,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,1,?,?)
        """, (*r, now, now))



# ── CRUD helpers ──────────────────────────────────────────────────────────
def get_setting(key, default=""):
    conn = get_conn()
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row["value"] if row else default

def set_setting(key, value):
    conn = get_conn()
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key, value))
    conn.commit(); conn.close()

def get_all_settings():
    conn = get_conn()
    rows = conn.execute("SELECT key,value FROM settings").fetchall()
    conn.close()
    return {r["key"]: r["value"] for r in rows}

def get_active_environment():
    conn = get_conn()
    row = conn.execute("SELECT * FROM environments WHERE is_active=1 LIMIT 1").fetchone()
    conn.close()
    return dict(row) if row else None

def get_all_environments():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM environments ORDER BY name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_environment(env_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM environments WHERE id=?", (env_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def save_environment(data, env_id=None):
    now = datetime.now().isoformat()
    conn = get_conn()
    if env_id:
        conn.execute("""
            UPDATE environments SET name=?,instance_url=?,username=?,password=?,
            table_name=?,page_size=?,timeout=?,start_date=?,start_time=?,
            end_date=?,end_time=?,level=?,updated_at=? WHERE id=?
        """, (data["name"],data["instance_url"],data["username"],data["password"],
              data.get("table_name","syslog"),int(data.get("page_size",100)),
              int(data.get("timeout",30)),data.get("start_date",""),
              data.get("start_time","00:00:00"),data.get("end_date",""),
              data.get("end_time","23:59:59"),data.get("level","2"),now,env_id))
    else:
        conn.execute("""
            INSERT INTO environments
            (name,instance_url,username,password,table_name,page_size,timeout,
             start_date,start_time,end_date,end_time,level,is_active,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,0,?,?)
        """, (data["name"],data["instance_url"],data["username"],data["password"],
              data.get("table_name","syslog"),int(data.get("page_size",100)),
              int(data.get("timeout",30)),data.get("start_date",""),
              data.get("start_time","00:00:00"),data.get("end_date",""),
              data.get("end_time","23:59:59"),data.get("level","2"),now,now))
    conn.commit(); conn.close()

def delete_environment(env_id):
    conn = get_conn()
    conn.execute("DELETE FROM environments WHERE id=?", (env_id,))
    conn.commit(); conn.close()

def set_active_environment(env_id):
    conn = get_conn()
    conn.execute("UPDATE environments SET is_active=0")
    conn.execute("UPDATE environments SET is_active=1 WHERE id=?", (env_id,))
    conn.commit(); conn.close()

def get_all_rules(search="", category="", severity="", active_only=False):
    conn = get_conn()
    q = "SELECT * FROM rules WHERE 1=1"
    params = []
    if search:
        q += " AND (name LIKE ? OR description LIKE ? OR pattern LIKE ? OR rule_id LIKE ?)"
        s = f"%{search}%"
        params += [s,s,s,s]
    if category:
        q += " AND category=?"
        params.append(category)
    if severity:
        q += " AND severity=?"
        params.append(severity)
    if active_only:
        q += " AND is_active=1"
    q += " ORDER BY rule_id"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_rule(rule_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM rules WHERE id=?", (rule_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def save_rule(data, rule_id=None):
    now = datetime.now().isoformat()
    conn = get_conn()
    if rule_id:
        conn.execute("""
            UPDATE rules SET rule_id=?,name=?,description=?,pattern=?,category=?,
            severity=?,source_hint=?,recommendation=?,affected_component=?,
            tags=?,is_active=?,updated_at=? WHERE id=?
        """, (data["rule_id"],data["name"],data.get("description",""),
              data["pattern"],data.get("category",""),data.get("severity","Medium"),
              data.get("source_hint",""),data.get("recommendation",""),
              data.get("affected_component",""),
              json.dumps(data.get("tags",[])) if isinstance(data.get("tags"),list) else data.get("tags","[]"),
              1 if data.get("is_active",True) else 0,now,rule_id))
    else:
        conn.execute("""
            INSERT INTO rules
            (rule_id,name,description,pattern,category,severity,source_hint,
             recommendation,affected_component,tags,is_active,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (data["rule_id"],data["name"],data.get("description",""),
              data["pattern"],data.get("category",""),data.get("severity","Medium"),
              data.get("source_hint",""),data.get("recommendation",""),
              data.get("affected_component",""),
              json.dumps(data.get("tags",[])) if isinstance(data.get("tags"),list) else data.get("tags","[]"),
              1 if data.get("is_active",True) else 0,now,now))
    conn.commit(); conn.close()

def delete_rule(rule_id):
    conn = get_conn()
    conn.execute("DELETE FROM rules WHERE id=?", (rule_id,))
    conn.commit(); conn.close()

def toggle_rule(rule_id):
    conn = get_conn()
    conn.execute("UPDATE rules SET is_active=CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id=?", (rule_id,))
    conn.commit(); conn.close()

def get_rule_categories():
    conn = get_conn()
    rows = conn.execute("SELECT DISTINCT category FROM rules WHERE category!='' ORDER BY category").fetchall()
    conn.close()
    return [r["category"] for r in rows]

def save_run(env, query, total, counts, status, xlsx="", html=""):
    now = datetime.now().isoformat()
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO analysis_runs
        (environment,start_date,start_time,end_date,end_time,level,
         total_records,critical_count,high_count,medium_count,low_count,
         status,output_xlsx,output_html,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (env,query["start_date"],query["start_time"],query["end_date"],
          query["end_time"],query["level"],total,
          counts.get("Critical",0),counts.get("High",0),
          counts.get("Medium",0),counts.get("Low",0),
          status,xlsx,html,now))
    run_id = cur.lastrowid
    conn.commit(); conn.close()
    return run_id

def update_run(run_id, total, counts, status, xlsx="", html=""):
    conn = get_conn()
    conn.execute("""
        UPDATE analysis_runs SET total_records=?,critical_count=?,high_count=?,
        medium_count=?,low_count=?,status=?,output_xlsx=?,output_html=? WHERE id=?
    """, (total,counts.get("Critical",0),counts.get("High",0),
          counts.get("Medium",0),counts.get("Low",0),status,xlsx,html,run_id))
    conn.commit(); conn.close()

def save_results(run_id, results):
    conn = get_conn()
    conn.execute("DELETE FROM analysis_results WHERE run_id=?", (run_id,))
    for r in results:
        conn.execute("""
            INSERT INTO analysis_results
            (run_id,created,level,message,source,node,job_name,user_name,thread,
             transaction_id,script_record,script_table,scope,page_name,record_sys_id,
             root_cause,matched_rule_ids,matched_rule_names,categories,highest_severity,
             recommendations,affected_components,rules_matched_count)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (run_id,r.get("created",""),r.get("level",""),r.get("message",""),
              r.get("source",""),r.get("node",""),r.get("job_name",""),
              r.get("user",""),r.get("thread",""),r.get("transaction_id",""),
              r.get("script_record",""),r.get("script_table",""),r.get("scope",""),
              r.get("page_name",""),r.get("record_sys_id",""),r.get("root_cause",""),
              r.get("matched_rule_ids",""),r.get("matched_rule_names",""),
              r.get("categories",""),r.get("highest_severity",""),
              r.get("recommendations",""),r.get("affected_components",""),
              r.get("rules_matched_count",0)))
    conn.commit(); conn.close()

def get_all_runs():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM analysis_runs ORDER BY id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_run(run_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM analysis_runs WHERE id=?", (run_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def get_run_results(run_id, search="", severity="", category="", node="", limit=500, offset=0):
    conn = get_conn()
    q = "SELECT * FROM analysis_results WHERE run_id=?"
    params = [run_id]
    if search:
        s = f"%{search}%"
        q += " AND (message LIKE ? OR node LIKE ? OR job_name LIKE ? OR user_name LIKE ? OR matched_rule_names LIKE ? OR categories LIKE ? OR root_cause LIKE ?)"
        params += [s,s,s,s,s,s,s]
    if severity:
        q += " AND highest_severity=?"
        params.append(severity)
    if category:
        q += " AND categories LIKE ?"
        params.append(f"%{category}%")
    if node:
        q += " AND node LIKE ?"
        params.append(f"%{node}%")
    total_q = q.replace("SELECT *","SELECT COUNT(*)")
    total = conn.execute(total_q, params).fetchone()[0]
    q += f" ORDER BY id LIMIT {limit} OFFSET {offset}"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows], total

def delete_run(run_id):
    conn = get_conn()
    conn.execute("DELETE FROM analysis_runs WHERE id=?", (run_id,))
    conn.commit(); conn.close()
''')

    # ── analyzer.py ───────────────────────────────────────────────────────
    create_file(f"{base}/analyzer.py", '''\
"""Log Analyzer — applies active rules from SQLite to syslog records."""
import re
import json
import logging
from typing import List, Dict, Any
from database import get_all_rules

logger = logging.getLogger(__name__)
SEV_ORDER = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1, "": 0}


class LogAnalyzer:
    def __init__(self):
        self.rules = get_all_rules(active_only=True)
        for rule in self.rules:
            try:
                rule["_cp"] = re.compile(rule["pattern"], re.IGNORECASE | re.DOTALL)
            except re.error as e:
                logger.warning(f"Bad regex in {rule.get('rule_id','?')}: {e}")
                rule["_cp"] = None
        logger.info(f"Analyzer loaded {len(self.rules)} active rules from DB.")

    def _get(self, record: Dict, field: str) -> str:
        v = record.get(field, "")
        if isinstance(v, dict):
            return v.get("value", "") or v.get("display_value", "")
        return str(v) if v else ""

    def _parse_ctx(self, ctx_str: str) -> Dict:
        try:
            return json.loads(ctx_str) if ctx_str else {}
        except Exception:
            return {}

    def analyze(self, records: List[Dict]) -> List[Dict]:
        results = []
        for rec in records:
            msg     = self._get(rec, "message")
            source  = self._get(rec, "source")
            ctx_str = self._get(rec, "context_map")
            ctx     = self._parse_ctx(ctx_str)
            full    = f"{msg} {ctx_str}"

            matched = [r for r in self.rules if r["_cp"] and r["_cp"].search(full)]

            parts = []
            for k, label in [("job_name","Job"),("_script_table","Script Table"),
                              ("_script_record","Script Record"),("_scope","Scope")]:
                v = ctx.get(k, "")
                if v: parts.append(f"{label}: {v}")
            if source: parts.append(f"Source: {source}")
            root_cause = " | ".join(parts) or "Unknown"

            top_sev = max(
                (r.get("severity","") for r in matched),
                key=lambda s: SEV_ORDER.get(s,0), default=""
            )

            node = ctx.get("_system_id","")
            short_node = node.split(":")[-1] if ":" in node else node

            results.append({
                "created":             self._get(rec,"sys_created_on"),
                "level":               self._get(rec,"level"),
                "message":             msg[:1000],
                "source":              source,
                "node":                short_node,
                "job_name":            ctx.get("job_name",""),
                "user":                ctx.get("_logged_in_user",""),
                "thread":              ctx.get("_thread_name",""),
                "transaction_id":      ctx.get("_txid",""),
                "script_record":       ctx.get("_script_record",""),
                "script_table":        ctx.get("_script_table",""),
                "scope":               ctx.get("_scope",""),
                "page_name":           ctx.get("_page_name",""),
                "record_sys_id":       ctx.get("record_sys_id",""),
                "root_cause":          root_cause,
                "matched_rule_ids":    ", ".join(r["rule_id"] for r in matched),
                "matched_rule_names":  ", ".join(r["name"] for r in matched),
                "categories":          ", ".join(sorted({r.get("category","") for r in matched})),
                "highest_severity":    top_sev,
                "recommendations":     " | ".join({r.get("recommendation","") for r in matched}),
                "affected_components": ", ".join(sorted({r.get("affected_component","") for r in matched})),
                "rules_matched_count": len(matched),
            })
        logger.info(f"Analysis complete: {len(results)} results.")
        return results
''')

    # ── sn_client.py ──────────────────────────────────────────────────────
    create_file(f"{base}/sn_client.py", '''\
"""ServiceNow Table API Client — persistent session."""
import requests
import logging
from typing import List, Dict, Any
logger = logging.getLogger(__name__)

class ServiceNowClient:
    def __init__(self, env: Dict[str, Any]):
        self.base      = env["instance_url"].rstrip("/")
        self.table     = env.get("table_name","syslog")
        self.page_size = int(env.get("page_size",100))
        self.timeout   = int(env.get("timeout",30))
        self.session   = requests.Session()
        self.session.auth = (env["username"], env["password"])
        self.session.headers.update({"Accept":"application/json","Content-Type":"application/json"})
        logger.info(f"SN Client → {self.base}")

    def _query(self, q):
        return (f"sys_created_on>javascript:gs.dateGenerate(\'{q['start_date']}\',\'{q['start_time']}\')"
                f"^level={q['level']}"
                f"^sys_created_on<javascript:gs.dateGenerate(\'{q['end_date']}\',\'{q['end_time']}\')")

    def fetch_records(self, qcfg: Dict) -> List[Dict]:
        url, query, all_r, offset = f"{self.base}/api/now/table/{self.table}", self._query(qcfg), [], 0
        logger.info(f"Query: {query}")
        while True:
            r = self.session.get(url, params={
                "sysparm_query":query,"sysparm_limit":self.page_size,
                "sysparm_offset":offset,
                "sysparm_fields":"sys_created_on,level,message,source,context_map,sys_created_by,sys_class_name,sequence,source_application_family,source_package",
                "sysparm_display_value":"false"
            }, timeout=self.timeout)
            r.raise_for_status()
            recs = r.json().get("result",[])
            if not recs: break
            all_r.extend(recs)
            logger.info(f"Fetched {len(recs)} @ offset {offset}. Total: {len(all_r)}")
            if len(recs) < self.page_size: break
            offset += self.page_size
        return all_r

    def close(self): self.session.close()
''')

    # ── reporter.py ───────────────────────────────────────────────────────
    create_file(f"{base}/reporter.py", '''\
"""Reporter — XLSX output only (HTML is served by Flask)."""
import os, logging
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
logger = logging.getLogger(__name__)

def _to_argb(h):
    h = h.lstrip("#").upper()
    if len(h)==3: h="".join(c*2 for c in h)
    if len(h)==6: h="FF"+h
    return h if len(h)==8 else "FF000000"

SEV_HEX={"Critical":("#C0392B","#FFFFFF"),"High":("#E67E22","#FFFFFF"),
          "Medium":("#F1C40F","#1A1A1A"),"Low":("#27AE60","#FFFFFF"),"":("#95A5A6","#FFFFFF")}

def _fill(c): a=_to_argb(c); return PatternFill(start_color=a,end_color=a,fill_type="solid")
def _font(c,bold=False,size=10): return Font(color=_to_argb(c),bold=bold,size=size)
def _bdr():
    t=Side(style="thin",color="FFCCCCCC")
    return Border(left=t,right=t,top=t,bottom=t)

def generate_xlsx(results: List[Dict], folder: str, prefix: str, env: str) -> str:
    os.makedirs(folder, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(folder, f"{prefix}{env}_{ts}.xlsx")
    wb   = openpyxl.Workbook()

    def hrow(ws, hdrs, row=1):
        for c,h in enumerate(hdrs,1):
            cl=ws.cell(row=row,column=c,value=h)
            cl.fill=_fill("#1F4E79"); cl.font=_font("#FFFFFF",bold=True,size=10)
            cl.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cl.border=_bdr()

    def fit(ws,mn=10,mx=55):
        for col in ws.columns:
            L=get_column_letter(col[0].column)
            w=max((len(str(c.value)) for c in col if c.value),default=mn)
            ws.column_dimensions[L].width=min(max(w+2,mn),mx)

    ws=wb.active; ws.title="Detailed Analysis"; ws.freeze_panes="A2"
    hdrs=["Timestamp","Level","Node","Job Name","User","Thread","Tx ID",
          "Script Table","Script Record","Scope","Page","Source",
          "Error Message","Root Cause","Rule IDs","Rule Names",
          "Categories","Severity","Recommendations","Components","# Rules"]
    hrow(ws,hdrs)
    for ri,r in enumerate(results,2):
        sev=r.get("highest_severity","")
        bg,fg=SEV_HEX.get(sev,("#FFFFFF","#000000")) if sev in SEV_HEX else ("#EBF3FB" if ri%2==0 else "#FFFFFF","#000000")
        rf=_fill(bg); rft=_font(fg,size=9)
        vals=[r.get("created",""),r.get("level",""),r.get("node",""),r.get("job_name",""),
              r.get("user",""),r.get("thread",""),r.get("transaction_id",""),
              r.get("script_table",""),r.get("script_record",""),r.get("scope",""),
              r.get("page_name",""),r.get("source",""),r.get("message","")[:400],
              r.get("root_cause",""),r.get("matched_rule_ids",""),r.get("matched_rule_names",""),
              r.get("categories",""),sev,r.get("recommendations","")[:300],
              r.get("affected_components",""),r.get("rules_matched_count",0)]
        for ci,v in enumerate(vals,1):
            cl=ws.cell(row=ri,column=ci,value=v)
            cl.fill=rf; cl.font=rft; cl.border=_bdr()
            cl.alignment=Alignment(vertical="top",wrap_text=True)
    fit(ws)

    for title,key,multi in [("By Category","categories",True),("By Node","node",False),("By Job","job_name",False)]:
        ws2=wb.create_sheet(title); ws2.freeze_panes="A2"
        hrow(ws2,["Group","Count","%"])
        cnt={}
        for r in results:
            vals2=r.get(key,"").split(", ") if multi else [r.get(key,"") or "Unknown"]
            for g in vals2:
                g=g.strip()
                if g: cnt[g]=cnt.get(g,0)+1
        tot=sum(cnt.values()) or 1
        for ri2,(g,c) in enumerate(sorted(cnt.items(),key=lambda x:-x[1]),2):
            rf2=_fill("#EBF3FB" if ri2%2==0 else "#FFFFFF")
            for ci2,v in enumerate([g,c,f"{c/tot*100:.1f}%"],1):
                cl2=ws2.cell(row=ri2,column=ci2,value=v)
                cl2.fill=rf2; cl2.font=_font("#000000",size=10)
                cl2.border=_bdr(); cl2.alignment=Alignment(vertical="top",wrap_text=True)
        fit(ws2)

    wb.save(path)
    logger.info(f"XLSX: {path}")
    return path
''')

    # ── app.py ────────────────────────────────────────────────────────────
    create_file(f"{base}/app.py", '''\
"""
Flask application — ServiceNow Syslog Analyzer
Routes:
  GET  /                        → Dashboard (run list)
  GET  /environments            → Environment list
  POST /environments            → Create environment
  GET  /environments/<id>       → Get environment JSON
  PUT  /environments/<id>       → Update environment
  DELETE /environments/<id>     → Delete environment
  POST /environments/<id>/activate → Set active
  GET  /rules                   → Rules management page
  GET  /api/rules               → Rules JSON (search/filter)
  POST /api/rules               → Create rule
  GET  /api/rules/<id>          → Get rule JSON
  PUT  /api/rules/<id>          → Update rule
  DELETE /api/rules/<id>        → Delete rule
  POST /api/rules/<id>/toggle   → Toggle active
  POST /run                     → Start analysis run
  GET  /run/<id>                → View run results (HTML report)
  GET  /api/run/<id>/results    → Results JSON (search/filter/page)
  GET  /api/run/<id>/summary    → Summary JSON
  DELETE /api/run/<id>          → Delete run
  GET  /download/<id>/xlsx      → Download XLSX
  GET  /settings                → Settings page
  POST /settings                → Save settings
"""
import os
import json
import logging
import threading
from datetime import datetime
from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, send_file, flash)

from database import (
    init_db, get_all_settings, set_setting, get_setting,
    get_all_environments, get_environment, save_environment,
    delete_environment, set_active_environment, get_active_environment,
    get_all_rules, get_rule, save_rule, delete_rule, toggle_rule,
    get_rule_categories, save_run, update_run, save_results,
    get_all_runs, get_run, get_run_results, delete_run
)
from analyzer  import LogAnalyzer
from sn_client import ServiceNowClient
from reporter  import generate_xlsx

app = Flask(__name__)
app.secret_key = "sn_syslog_analyzer_secret_2026"

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s")
logger = logging.getLogger(__name__)

# ── Init DB on startup ────────────────────────────────────────────────────
with app.app_context():
    init_db()

# ── Background run tracker ────────────────────────────────────────────────
run_status = {}  # run_id → {"status":"running","progress":"..."}


def _do_run(run_id, env, query, output_folder, output_prefix):
    """Background thread: fetch → analyse → save → report."""
    run_status[run_id] = {"status":"running","progress":"Connecting to ServiceNow…"}
    try:
        client = ServiceNowClient(env)
        run_status[run_id]["progress"] = "Fetching records…"
        records = client.fetch_records(query)
        client.close()

        run_status[run_id]["progress"] = f"Analysing {len(records)} records…"
        analyzer = LogAnalyzer()
        results  = analyzer.analyze(records)

        sc = {}
        for r in results:
            s = r.get("highest_severity","Unknown") or "Unknown"
            sc[s] = sc.get(s,0)+1

        run_status[run_id]["progress"] = "Saving results to DB…"
        save_results(run_id, results)

        run_status[run_id]["progress"] = "Generating XLSX…"
        xlsx_path = generate_xlsx(results, output_folder, output_prefix, env["name"])

        update_run(run_id, len(results), sc, "completed", xlsx=xlsx_path)
        run_status[run_id] = {"status":"completed","progress":"Done"}
        logger.info(f"Run {run_id} completed. {len(results)} results.")
    except Exception as e:
        logger.error(f"Run {run_id} failed: {e}", exc_info=True)
        update_run(run_id, 0, {}, "failed")
        run_status[run_id] = {"status":"failed","progress":str(e)}


# ── Dashboard ─────────────────────────────────────────────────────────────
@app.route("/")
def dashboard():
    runs   = get_all_runs()
    active = get_active_environment()
    return render_template("dashboard.html", runs=runs, active_env=active)

# ── Run analysis ──────────────────────────────────────────────────────────
@app.route("/run", methods=["POST"])
def start_run():
    env = get_active_environment()
    if not env:
        flash("No active environment. Please configure one first.", "error")
        return redirect(url_for("environments_page"))

    # Allow query override from form
    query = {
        "start_date": request.form.get("start_date", env.get("start_date","")),
        "start_time": request.form.get("start_time", env.get("start_time","00:00:00")),
        "end_date":   request.form.get("end_date",   env.get("end_date","")),
        "end_time":   request.form.get("end_time",   env.get("end_time","23:59:59")),
        "level":      request.form.get("level",      env.get("level","2")),
    }

    settings = get_all_settings()
    run_id = save_run(env["name"], query, 0, {}, "running")
    t = threading.Thread(
        target=_do_run,
        args=(run_id, env, query,
              settings.get("output_folder","output"),
              settings.get("output_prefix","output_error_")),
        daemon=True
    )
    t.start()
    return redirect(url_for("view_run", run_id=run_id))

@app.route("/api/run/<int:run_id>/status")
def run_status_api(run_id):
    st = run_status.get(run_id, {"status":"unknown","progress":""})
    run = get_run(run_id)
    if run: st["db_status"] = run["status"]
    return jsonify(st)

@app.route("/run/<int:run_id>")
def view_run(run_id):
    run = get_run(run_id)
    if not run:
        flash("Run not found.", "error")
        return redirect(url_for("dashboard"))
    return render_template("run_detail.html", run=run)

@app.route("/api/run/<int:run_id>/results")
def run_results_api(run_id):
    search   = request.args.get("search","")
    severity = request.args.get("severity","")
    category = request.args.get("category","")
    node     = request.args.get("node","")
    limit    = int(request.args.get("limit",200))
    offset   = int(request.args.get("offset",0))
    rows, total = get_run_results(run_id, search, severity, category, node, limit, offset)
    return jsonify({"results":rows,"total":total,"limit":limit,"offset":offset})

@app.route("/api/run/<int:run_id>/summary")
def run_summary_api(run_id):
    rows, _ = get_run_results(run_id, limit=9999)
    sc={};nc={};jc={};cc={}
    for r in rows:
        s=r.get("highest_severity","Unknown") or "Unknown"
        sc[s]=sc.get(s,0)+1
        n=r.get("node","Unknown") or "Unknown"
        nc[n]=nc.get(n,0)+1
        j=r.get("job_name","Unknown") or "Unknown"
        jc[j]=jc.get(j,0)+1
        for cat in r.get("categories","").split(", "):
            cat=cat.strip()
            if cat: cc[cat]=cc.get(cat,0)+1
    return jsonify({"severity":sc,"node":nc,"job":jc,"category":cc,"total":len(rows)})

@app.route("/api/run/<int:run_id>", methods=["DELETE"])
def delete_run_api(run_id):
    delete_run(run_id)
    return jsonify({"ok":True})

@app.route("/download/<int:run_id>/xlsx")
def download_xlsx(run_id):
    run = get_run(run_id)
    if not run or not run.get("output_xlsx") or not os.path.exists(run["output_xlsx"]):
        return "File not found", 404
    return send_file(run["output_xlsx"], as_attachment=True,
                     download_name=os.path.basename(run["output_xlsx"]))

# ── Environments ──────────────────────────────────────────────────────────
@app.route("/environments")
def environments_page():
    envs = get_all_environments()
    return render_template("environments.html", envs=envs)

@app.route("/api/environments", methods=["GET"])
def api_get_envs():
    return jsonify(get_all_environments())

@app.route("/api/environments", methods=["POST"])
def api_create_env():
    data = request.json
    save_environment(data)
    return jsonify({"ok":True})

@app.route("/api/environments/<int:env_id>", methods=["GET"])
def api_get_env(env_id):
    env = get_environment(env_id)
    if not env: return jsonify({"error":"Not found"}),404
    return jsonify(env)

@app.route("/api/environments/<int:env_id>", methods=["PUT"])
def api_update_env(env_id):
    data = request.json
    save_environment(data, env_id)
    return jsonify({"ok":True})

@app.route("/api/environments/<int:env_id>", methods=["DELETE"])
def api_delete_env(env_id):
    delete_environment(env_id)
    return jsonify({"ok":True})

@app.route("/api/environments/<int:env_id>/activate", methods=["POST"])
def api_activate_env(env_id):
    set_active_environment(env_id)
    return jsonify({"ok":True})

# ── Rules ─────────────────────────────────────────────────────────────────
@app.route("/rules")
def rules_page():
    return render_template("rules.html")

@app.route("/api/rules", methods=["GET"])
def api_get_rules():
    search   = request.args.get("search","")
    category = request.args.get("category","")
    severity = request.args.get("severity","")
    active   = request.args.get("active_only","") == "1"
    rules    = get_all_rules(search, category, severity, active)
    cats     = get_rule_categories()
    return jsonify({"rules":rules,"categories":cats,"total":len(rules)})

@app.route("/api/rules", methods=["POST"])
def api_create_rule():
    data = request.json
    save_rule(data)
    return jsonify({"ok":True})

@app.route("/api/rules/<int:rule_id>", methods=["GET"])
def api_get_rule(rule_id):
    rule = get_rule(rule_id)
    if not rule: return jsonify({"error":"Not found"}),404
    return jsonify(rule)

@app.route("/api/rules/<int:rule_id>", methods=["PUT"])
def api_update_rule(rule_id):
    data = request.json
    save_rule(data, rule_id)
    return jsonify({"ok":True})

@app.route("/api/rules/<int:rule_id>", methods=["DELETE"])
def api_delete_rule(rule_id):
    delete_rule(rule_id)
    return jsonify({"ok":True})

@app.route("/api/rules/<int:rule_id>/toggle", methods=["POST"])
def api_toggle_rule(rule_id):
    toggle_rule(rule_id)
    return jsonify({"ok":True})

# ── Settings ──────────────────────────────────────────────────────────────
@app.route("/settings", methods=["GET","POST"])
def settings_page():
    if request.method == "POST":
        for key in ["output_folder","output_prefix","output_formats","log_level"]:
            val = request.form.get(key,"")
            if val: set_setting(key, val)
        flash("Settings saved.", "success")
        return redirect(url_for("settings_page"))
    return render_template("settings.html", settings=get_all_settings())

if __name__ == "__main__":
    app.run(debug=True, port=5000)
''')

    # ── templates/base.html ───────────────────────────────────────────────
    create_file(f"{base}/templates/base.html", """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{% block title %}SN Syslog Analyzer{% endblock %}</title>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#0f1117;--surface:#1a1d27;--surface2:#22263a;--surface3:#2a2f47;
  --accent:#4f8ef7;--accent2:#7c5cfc;--muted:#8892b0;--text:#e2e8f0;--text2:#a0aec0;
  --border:#2d3561;--radius:10px;--shadow:0 4px 24px #0006;
  --critical:#e74c3c;--high:#e67e22;--medium:#f1c40f;--low:#27ae60;--ok:#27ae60;
}
body{font-family:'Segoe UI',system-ui,sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
a{color:var(--accent);text-decoration:none}
a:hover{text-decoration:underline}
/* Topbar */
.topbar{background:linear-gradient(135deg,#1a1d27,#22263a);border-bottom:1px solid var(--border);
  padding:0 24px;display:flex;align-items:center;height:54px;position:sticky;top:0;z-index:200;
  box-shadow:0 2px 20px #0008}
.topbar-logo{font-size:1.25em;font-weight:800;background:linear-gradient(90deg,var(--accent),var(--accent2));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;white-space:nowrap}
.topbar-nav{display:flex;align-items:center;gap:4px;margin-left:28px;flex:1}
.nav-link{padding:6px 14px;border-radius:8px;color:var(--text2);font-size:0.87em;
  font-weight:500;transition:all .16s;white-space:nowrap}
.nav-link:hover{background:var(--surface3);color:var(--text);text-decoration:none}
.nav-link.active{background:linear-gradient(90deg,#4f8ef718,#7c5cfc18);color:var(--accent)}
.topbar-right{margin-left:auto;display:flex;align-items:center;gap:10px}
.env-pill{background:var(--surface3);border:1px solid var(--border);border-radius:20px;
  padding:4px 12px;font-size:0.75em;font-weight:700;color:var(--accent)}
/* Layout */
.page{max-width:1600px;margin:0 auto;padding:24px 20px}
/* Cards */
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:20px}
.card-title{font-size:0.8em;font-weight:700;color:var(--muted);text-transform:uppercase;
  letter-spacing:1px;margin-bottom:14px}
/* Buttons */
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 16px;border-radius:8px;
  border:none;cursor:pointer;font-size:0.86em;font-weight:600;transition:all .16s;white-space:nowrap}
.btn-primary{background:linear-gradient(135deg,var(--accent),var(--accent2));color:#fff}
.btn-primary:hover{opacity:.9;transform:translateY(-1px)}
.btn-danger{background:#e74c3c22;color:#e74c3c;border:1px solid #e74c3c44}
.btn-danger:hover{background:#e74c3c33}
.btn-success{background:#27ae6022;color:#27ae60;border:1px solid #27ae6044}
.btn-success:hover{background:#27ae6033}
.btn-ghost{background:var(--surface2);color:var(--text2);border:1px solid var(--border)}
.btn-ghost:hover{background:var(--surface3);color:var(--text)}
.btn-sm{padding:5px 10px;font-size:0.78em}
/* Form */
.form-group{margin-bottom:14px}
.form-label{display:block;font-size:0.75em;font-weight:600;color:var(--muted);
  text-transform:uppercase;letter-spacing:.8px;margin-bottom:5px}
.form-control{width:100%;background:var(--surface2);border:1px solid var(--border);
  border-radius:8px;padding:8px 12px;color:var(--text);font-size:0.87em;outline:none;transition:border .2s}
.form-control:focus{border-color:var(--accent)}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.form-row-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px}
/* Table */
.tbl{width:100%;border-collapse:separate;border-spacing:0;font-size:0.83em}
.tbl thead th{background:var(--surface2);color:var(--muted);font-weight:700;font-size:0.76em;
  text-transform:uppercase;letter-spacing:.8px;padding:10px 12px;
  border-bottom:2px solid var(--border);white-space:nowrap;position:sticky;top:0;z-index:10}
.tbl tbody tr{transition:background .12s;cursor:pointer}
.tbl tbody tr:hover{background:var(--surface3)}
.tbl td{padding:9px 12px;border-bottom:1px solid var(--border);color:var(--text2);
  max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;vertical-align:middle}
.tbl tbody tr:hover td{color:var(--text)}
/* Severity badge */
.sev{display:inline-flex;align-items:center;gap:3px;padding:2px 8px;border-radius:20px;
  font-size:0.72em;font-weight:700;white-space:nowrap}
.sev::before{content:"●";font-size:.65em}
.sev.Critical{background:#e74c3c22;color:#e74c3c;border:1px solid #e74c3c55}
.sev.High{background:#e67e2222;color:#e67e22;border:1px solid #e67e2255}
.sev.Medium{background:#f1c40f22;color:#f1c40f;border:1px solid #f1c40f55}
.sev.Low{background:#27ae6022;color:#27ae60;border:1px solid #27ae6055}
.sev.Unknown{background:#95a5a622;color:#95a5a6;border:1px solid #95a5a655}
/* Status badge */
.status{display:inline-block;padding:2px 9px;border-radius:20px;font-size:0.72em;font-weight:700}
.status.completed{background:#27ae6022;color:#27ae60;border:1px solid #27ae6044}
.status.running{background:#4f8ef722;color:#4f8ef7;border:1px solid #4f8ef744}
.status.failed{background:#e74c3c22;color:#e74c3c;border:1px solid #e74c3c44}
.status.pending{background:#95a5a622;color:#95a5a6;border:1px solid #95a5a644}
/* Modal */
.modal-overlay{position:fixed;inset:0;background:#0009;z-index:500;display:flex;
  align-items:center;justify-content:center;opacity:0;pointer-events:none;transition:opacity .2s}
.modal-overlay.open{opacity:1;pointer-events:all}
.modal{background:var(--surface);border:1px solid var(--border);border-radius:14px;
  padding:28px;width:min(600px,95vw);max-height:90vh;overflow-y:auto;
  box-shadow:0 20px 60px #0008;transform:translateY(20px);transition:transform .2s}
.modal-overlay.open .modal{transform:translateY(0)}
.modal-title{font-size:1.05em;font-weight:700;margin-bottom:20px;
  padding-bottom:12px;border-bottom:1px solid var(--border)}
.modal-footer{display:flex;gap:10px;justify-content:flex-end;margin-top:20px;
  padding-top:14px;border-top:1px solid var(--border)}
/* Alert */
.alert{padding:10px 16px;border-radius:8px;font-size:0.86em;margin-bottom:14px}
.alert-success{background:#27ae6022;color:#27ae60;border:1px solid #27ae6044}
.alert-error{background:#e74c3c22;color:#e74c3c;border:1px solid #e74c3c44}
/* Misc */
.tag{display:inline-block;background:var(--surface3);border:1px solid var(--border);
  border-radius:4px;padding:2px 7px;font-size:0.7em;color:var(--muted);margin:2px}
.empty{text-align:center;padding:50px 20px;color:var(--muted)}
.empty-icon{font-size:2.5em;margin-bottom:10px}
.flex{display:flex;align-items:center;gap:10px}
.flex-1{flex:1}
.gap-2{gap:8px}
.mt-1{margin-top:8px}
.mt-2{margin-top:16px}
.mb-2{margin-bottom:16px}
.text-muted{color:var(--muted);font-size:0.85em}
.text-right{text-align:right}
.bold{font-weight:700}
.w-full{width:100%}
/* Search bar */
.search-wrap{position:relative}
.search-wrap input{padding-left:32px}
.search-wrap .si{position:absolute;left:10px;top:50%;transform:translateY(-50%);color:var(--muted)}
/* Scrollbar */
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:var(--surface)}
::-webkit-scrollbar-thumb{background:var(--border);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--muted)}
/* Toggle switch */
.toggle{position:relative;display:inline-block;width:38px;height:20px}
.toggle input{opacity:0;width:0;height:0}
.toggle-slider{position:absolute;inset:0;background:var(--surface3);border-radius:20px;
  cursor:pointer;transition:.2s}
.toggle-slider::before{content:"";position:absolute;width:14px;height:14px;left:3px;bottom:3px;
  background:#fff;border-radius:50%;transition:.2s}
.toggle input:checked+.toggle-slider{background:var(--accent)}
.toggle input:checked+.toggle-slider::before{transform:translateX(18px)}
</style>
{% block extra_style %}{% endblock %}
</head>
<body>
<div class="topbar">
  <span class="topbar-logo">⚡ SN Syslog Analyzer</span>
  <nav class="topbar-nav">
    <a href="/" class="nav-link {% if request.path=='/' %}active{% endif %}">🏠 Dashboard</a>
    <a href="/environments" class="nav-link {% if '/environments' in request.path %}active{% endif %}">🌐 Environments</a>
    <a href="/rules" class="nav-link {% if '/rules' in request.path %}active{% endif %}">📋 Rules</a>
    <a href="/settings" class="nav-link {% if '/settings' in request.path %}active{% endif %}">⚙ Settings</a>
  </nav>
  <div class="topbar-right">
    {% if active_env %}
    <span class="env-pill">{{ active_env.name }}</span>
    {% endif %}
  </div>
</div>

{% with messages = get_flashed_messages(with_categories=true) %}
{% if messages %}
<div style="padding:10px 20px">
{% for cat,msg in messages %}
<div class="alert alert-{{ 'success' if cat=='success' else 'error' }}">{{ msg }}</div>
{% endfor %}
</div>
{% endif %}
{% endwith %}

{% block content %}{% endblock %}

{% block scripts %}{% endblock %}
</body>
</html>
""")

    # ── templates/dashboard.html ──────────────────────────────────────────
    create_file(f"{base}/templates/dashboard.html", """\
{% extends "base.html" %}
{% block title %}Dashboard — SN Syslog Analyzer{% endblock %}
{% block content %}
<div class="page">

  <!-- Run form -->
  <div class="card mb-2">
    <div class="card-title">▶ Start New Analysis Run</div>
    {% if active_env %}
    <form method="POST" action="/run" id="runForm">
      <div class="form-row-3">
        <div class="form-group">
          <label class="form-label">Start Date</label>
          <input class="form-control" type="date" name="start_date" value="{{ active_env.start_date }}" required/>
        </div>
        <div class="form-group">
          <label class="form-label">Start Time</label>
          <input class="form-control" type="time" name="start_time" value="{{ active_env.start_time }}" step="1" required/>
        </div>
        <div class="form-group">
          <label class="form-label">Level</label>
          <select class="form-control" name="level">
            <option value="2" {% if active_env.level=='2' %}selected{% endif %}>2 — Error</option>
            <option value="1" {% if active_env.level=='1' %}selected{% endif %}>1 — Warning</option>
            <option value="0" {% if active_env.level=='0' %}selected{% endif %}>0 — Info</option>
          </select>
        </div>
        <div class="form-group">
          <label class="form-label">End Date</label>
          <input class="form-control" type="date" name="end_date" value="{{ active_env.end_date }}" required/>
        </div>
        <div class="form-group">
          <label class="form-label">End Time</label>
          <input class="form-control" type="time" name="end_time" value="{{ active_env.end_time }}" step="1" required/>
        </div>
        <div class="form-group" style="display:flex;align-items:flex-end">
          <button class="btn btn-primary w-full" type="submit">⚡ Run Analysis</button>
        </div>
      </div>
      <div class="text-muted">Active environment: <strong>{{ active_env.name }}</strong> — {{ active_env.instance_url }}</div>
    </form>
    {% else %}
    <div class="alert alert-error">No active environment. <a href="/environments">Configure one →</a></div>
    {% endif %}
  </div>

  <!-- Run history -->
  <div class="card">
    <div class="flex mb-2">
      <div class="card-title flex-1" style="margin-bottom:0">📊 Analysis History</div>
    </div>
    {% if runs %}
    <div style="overflow-x:auto">
    <table class="tbl">
      <thead><tr>
        <th>#</th><th>Environment</th><th>Query Window</th>
        <th>Total</th><th>Critical</th><th>High</th><th>Medium</th><th>Low</th>
        <th>Status</th><th>Created</th><th>Actions</th>
      </tr></thead>
      <tbody>
      {% for r in runs %}
      <tr onclick="window.location='/run/{{ r.id }}'" style="cursor:pointer">
        <td>{{ r.id }}</td>
        <td><strong>{{ r.environment }}</strong></td>
        <td style="font-size:0.78em">{{ r.start_date }} {{ r.start_time }}<br>→ {{ r.end_date }} {{ r.end_time }}</td>
        <td><strong>{{ r.total_records }}</strong></td>
        <td><span class="sev Critical">{{ r.critical_count }}</span></td>
        <td><span class="sev High">{{ r.high_count }}</span></td>
        <td><span class="sev Medium">{{ r.medium_count }}</span></td>
        <td><span class="sev Low">{{ r.low_count }}</span></td>
        <td><span class="status {{ r.status }}">{{ r.status }}</span></td>
        <td style="font-size:0.78em;color:var(--muted)">{{ r.created_at[:16] }}</td>
        <td onclick="event.stopPropagation()">
          <div class="flex gap-2">
            <a href="/run/{{ r.id }}" class="btn btn-ghost btn-sm">View</a>
            {% if r.output_xlsx %}
            <a href="/download/{{ r.id }}/xlsx" class="btn btn-ghost btn-sm">⬇ XLSX</a>
            {% endif %}
            <button class="btn btn-danger btn-sm" onclick="deleteRun({{ r.id }})">🗑</button>
          </div>
        </td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    </div>
    {% else %}
    <div class="empty"><div class="empty-icon">📭</div><div>No analysis runs yet. Start one above.</div></div>
    {% endif %}
  </div>
</div>
{% endblock %}
{% block scripts %}
<script>
function deleteRun(id){
  if(!confirm('Delete this run and all its results?')) return;
  fetch('/api/run/'+id,{method:'DELETE'}).then(()=>location.reload());
}
</script>
{% endblock %}
""")

    # ── templates/run_detail.html ─────────────────────────────────────────
    create_file(f"{base}/templates/run_detail.html", """\
{% extends "base.html" %}
{% block title %}Run #{{ run.id }} — SN Syslog Analyzer{% endblock %}
{% block extra_style %}
.layout{display:flex;height:calc(100vh - 54px);overflow:hidden}
.sidebar{width:240px;min-width:200px;background:var(--surface);border-right:1px solid var(--border);
  display:flex;flex-direction:column;transition:width .25s;overflow:hidden}
.sidebar.collapsed{width:0;min-width:0}
.sidebar-inner{padding:12px 10px;flex:1;overflow-y:auto;display:flex;flex-direction:column;gap:3px}
.sidebar-title{font-size:0.67em;font-weight:700;color:var(--muted);text-transform:uppercase;
  letter-spacing:1.5px;padding:8px 4px 4px;white-space:nowrap}
.nav-btn{display:flex;align-items:center;gap:9px;padding:8px 11px;border-radius:8px;border:none;
  background:transparent;color:var(--text2);cursor:pointer;font-size:0.86em;width:100%;
  text-align:left;transition:all .16s;white-space:nowrap}
.nav-btn:hover{background:var(--surface3);color:var(--text)}
.nav-btn.active{background:linear-gradient(90deg,#4f8ef718,#7c5cfc18);color:var(--accent);border-left:3px solid var(--accent)}
.main{flex:1;display:flex;flex-direction:column;overflow:hidden}
.toolbar{background:var(--surface);border-bottom:1px solid var(--border);
  padding:9px 16px;display:flex;align-items:center;gap:9px;flex-wrap:wrap}
.content-area{flex:1;overflow:hidden;display:flex;flex-direction:column}
.cards-row{display:flex;gap:9px;padding:12px 16px;flex-wrap:wrap;
  background:var(--surface);border-bottom:1px solid var(--border)}
.stat-card{flex:1;min-width:90px;max-width:140px;background:var(--surface2);
  border:1px solid var(--border);border-radius:var(--radius);padding:11px 13px;
  cursor:pointer;transition:all .18s;position:relative;overflow:hidden}
.stat-card::before{content:"";position:absolute;top:0;left:0;right:0;height:3px}
.stat-card.all::before{background:linear-gradient(90deg,var(--accent),var(--accent2))}
.stat-card.Critical::before{background:var(--critical)}
.stat-card.High::before{background:var(--high)}
.stat-card.Medium::before{background:var(--medium)}
.stat-card.Low::before{background:var(--low)}
.stat-card:hover{transform:translateY(-2px);box-shadow:var(--shadow);border-color:var(--accent)}
.stat-card.active-card{border-color:var(--accent);background:var(--surface3)}
.stat-num{font-size:1.8em;font-weight:800;line-height:1}
.stat-lbl{font-size:0.7em;color:var(--muted);margin-top:3px;font-weight:600;text-transform:uppercase}
.stat-card.all .stat-num{color:var(--accent)}
.stat-card.Critical .stat-num{color:var(--critical)}
.stat-card.High .stat-num{color:var(--high)}
.stat-card.Medium .stat-num{color:var(--medium)}
.stat-card.Low .stat-num{color:var(--low)}
.view{display:none}.view.active{display:flex;flex-direction:column;flex:1;overflow:hidden}
.table-panel{flex:1;overflow:auto;padding:0 16px 16px}
.data-table{width:100%;border-collapse:separate;border-spacing:0;font-size:0.81em;margin-top:12px}
.data-table thead th{background:var(--surface2);color:var(--muted);font-weight:700;
  font-size:0.76em;text-transform:uppercase;letter-spacing:.8px;padding:9px 11px;
  border-bottom:2px solid var(--border);position:sticky;top:0;z-index:10;
  white-space:nowrap;cursor:pointer;user-select:none}
.data-table thead th:hover{color:var(--accent)}
.data-table thead th.sorted{color:var(--accent)}
.data-table tbody tr{cursor:pointer;transition:background .12s}
.data-table tbody tr:hover{background:var(--surface3)}
.data-table tbody tr.selected{background:linear-gradient(90deg,#4f8ef710,#7c5cfc10);
  outline:1px solid var(--accent);outline-offset:-1px}
.data-table td{padding:8px 11px;vertical-align:middle;max-width:230px;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--text2);
  border-bottom:1px solid var(--border)}
.data-table tbody tr:hover td,.data-table tbody tr.selected td{color:var(--text)}
.detail-panel{position:fixed;top:54px;right:0;bottom:0;width:480px;
  background:var(--surface);border-left:1px solid var(--border);
  display:flex;flex-direction:column;z-index:50;
  box-shadow:-8px 0 32px #0008;transition:transform .28s cubic-bezier(.4,0,.2,1)}
.detail-panel.hidden{transform:translateX(100%)}
.detail-panel.dragging{transition:none}
.dp-resize{position:absolute;left:0;top:0;bottom:0;width:5px;cursor:ew-resize;
  background:transparent;transition:background .2s}
.dp-resize:hover{background:var(--accent)}
.dp-header{padding:13px 16px;border-bottom:1px solid var(--border);
  display:flex;align-items:center;gap:9px;background:var(--surface2)}
.dp-title{font-weight:700;font-size:0.93em;flex:1}
.dp-close{background:none;border:none;color:var(--muted);cursor:pointer;
  font-size:1.15em;padding:2px 6px;border-radius:6px;transition:all .14s}
.dp-close:hover{background:var(--surface3);color:var(--text)}
.dp-drag{cursor:ew-resize;padding:0 7px;color:var(--muted);font-size:1.05em;user-select:none}
.dp-body{flex:1;overflow-y:auto;padding:16px}
.dp-section{margin-bottom:16px}
.dp-section-title{font-size:0.67em;font-weight:700;color:var(--muted);text-transform:uppercase;
  letter-spacing:1.5px;margin-bottom:7px;padding-bottom:5px;border-bottom:1px solid var(--border)}
.dp-grid{display:grid;grid-template-columns:1fr 1fr;gap:7px}
.dp-field{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:8px 10px}
.dp-field.full{grid-column:1/-1}
.dp-field-lbl{font-size:0.67em;color:var(--muted);font-weight:600;text-transform:uppercase;
  letter-spacing:.8px;margin-bottom:3px}
.dp-field-val{font-size:0.82em;color:var(--text);word-break:break-word;line-height:1.5}
.dp-msg{background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:10px;
  font-family:'Cascadia Code','Consolas',monospace;font-size:0.77em;color:#a8d8a8;
  white-space:pre-wrap;word-break:break-word;max-height:170px;overflow-y:auto;line-height:1.6}
.dp-rec{background:var(--surface2);border:1px solid var(--border);border-radius:8px;
  padding:8px 10px;margin-bottom:6px}
.dp-rec-id{font-size:0.69em;color:var(--accent);font-weight:700;margin-bottom:2px}
.dp-rec-name{font-size:0.82em;color:var(--text);font-weight:600}
.sum-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(270px,1fr));gap:13px;padding:16px}
.sum-card{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);padding:16px}
.sum-card h3{font-size:0.8em;color:var(--muted);text-transform:uppercase;letter-spacing:1px;margin-bottom:11px}
.sum-bar-row{display:flex;align-items:center;gap:8px;margin-bottom:6px;font-size:0.79em}
.sum-bar-lbl{width:110px;color:var(--text2);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.sum-bar-track{flex:1;background:var(--surface3);border-radius:4px;height:6px;overflow:hidden}
.sum-bar-fill{height:100%;border-radius:4px;transition:width .6s ease}
.sum-bar-cnt{width:30px;text-align:right;color:var(--muted)}
.spinner{display:inline-block;width:18px;height:18px;border:2px solid var(--border);
  border-top-color:var(--accent);border-radius:50%;animation:spin .7s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.progress-bar{height:4px;background:var(--surface3);border-radius:2px;overflow:hidden;margin-top:8px}
.progress-fill{height:100%;background:linear-gradient(90deg,var(--accent),var(--accent2));
  border-radius:2px;animation:prog 1.5s ease-in-out infinite alternate}
@keyframes prog{from{width:20%}to{width:90%}}
{% endblock %}
{% block content %}
<div class="layout">
<!-- Sidebar -->
<div class="sidebar" id="sidebar">
  <div class="sidebar-inner">
    <div class="sidebar-title">Run #{{ run.id }}</div>
    <button class="nav-btn active" onclick="switchView('detail',this)">📋 Detailed Results</button>
    <button class="nav-btn" onclick="switchView('summary',this)">📊 Summary</button>
    <div class="sidebar-title" style="margin-top:10px">Filter Severity</div>
    <button class="nav-btn" onclick="filterSev('',this)" style="border-left:3px solid var(--accent)">
      <span style="color:var(--accent)">●</span> All
    </button>
    <button class="nav-btn" onclick="filterSev('Critical',this)"><span style="color:var(--critical)">●</span> Critical</button>
    <button class="nav-btn" onclick="filterSev('High',this)"><span style="color:var(--high)">●</span> High</button>
    <button class="nav-btn" onclick="filterSev('Medium',this)"><span style="color:var(--medium)">●</span> Medium</button>
    <button class="nav-btn" onclick="filterSev('Low',this)"><span style="color:var(--low)">●</span> Low</button>
    <div class="sidebar-title" style="margin-top:10px">Actions</div>
    {% if run.output_xlsx %}
    <a href="/download/{{ run.id }}/xlsx" class="nav-btn">⬇ Download XLSX</a>
    {% endif %}
    <a href="/" class="nav-btn">← Back to Dashboard</a>
  </div>
</div>

<div class="main">
  <!-- Toolbar -->
  <div class="toolbar">
    <div class="search-wrap flex-1" style="max-width:400px">
      <span class="si">🔍</span>
      <input class="form-control" type="text" id="searchBox" placeholder="Search message, node, job, user, rule…" oninput="debounceSearch()"/>
    </div>
    <select class="form-control" id="catFilter" onchange="loadResults()" style="width:auto">
      <option value="">All Categories</option>
    </select>
    <select class="form-control" id="nodeFilter" onchange="loadResults()" style="width:auto">
      <option value="">All Nodes</option>
    </select>
    <button class="btn btn-ghost" onclick="clearFilters()">✕ Clear</button>
    <span id="countBadge" class="text-muted">—</span>
    <button class="btn btn-ghost" onclick="toggleDetail()" id="dpToggleBtn">⊞ Detail</button>
    <button class="btn btn-ghost" onclick="toggleSidebar()">☰</button>
  </div>

  <div class="content-area">
    <!-- Status / progress -->
    <div id="runningBanner" style="display:none;padding:10px 16px;background:var(--surface);
      border-bottom:1px solid var(--border);display:flex;align-items:center;gap:12px">
      <div class="spinner"></div>
      <div>
        <div style="font-size:0.87em;font-weight:600">Analysis in progress…</div>
        <div id="progressText" style="font-size:0.78em;color:var(--muted)"></div>
        <div class="progress-bar"><div class="progress-fill"></div></div>
      </div>
    </div>

    <!-- Stat cards -->
    <div class="cards-row" id="cardsRow">
      <div class="stat-card all active-card" onclick="filterSev('',null)" data-sev="">
        <div class="stat-num" id="cnt-all">{{ run.total_records }}</div><div class="stat-lbl">All</div>
      </div>
      <div class="stat-card Critical" onclick="filterSev('Critical',null)" data-sev="Critical">
        <div class="stat-num" id="cnt-Critical">{{ run.critical_count }}</div><div class="stat-lbl">Critical</div>
      </div>
      <div class="stat-card High" onclick="filterSev('High',null)" data-sev="High">
        <div class="stat-num" id="cnt-High">{{ run.high_count }}</div><div class="stat-lbl">High</div>
      </div>
      <div class="stat-card Medium" onclick="filterSev('Medium',null)" data-sev="Medium">
        <div class="stat-num" id="cnt-Medium">{{ run.medium_count }}</div><div class="stat-lbl">Medium</div>
      </div>
      <div class="stat-card Low" onclick="filterSev('Low',null)" data-sev="Low">
        <div class="stat-num" id="cnt-Low">{{ run.low_count }}</div><div class="stat-lbl">Low</div>
      </div>
    </div>

    <!-- Detail view -->
    <div class="view active" id="view-detail">
      <div class="table-panel">
        <table class="data-table">
          <thead><tr>
            <th onclick="sortTable('created')" data-col="created">Timestamp ⇅</th>
            <th onclick="sortTable('highest_severity')" data-col="highest_severity">Severity ⇅</th>
            <th onclick="sortTable('node')" data-col="node">Node ⇅</th>
            <th onclick="sortTable('job_name')" data-col="job_name">Job ⇅</th>
            <th onclick="sortTable('user_name')" data-col="user_name">User ⇅</th>
            <th onclick="sortTable('categories')" data-col="categories">Categories ⇅</th>
            <th onclick="sortTable('message')" data-col="message">Error Message ⇅</th>
            <th onclick="sortTable('rules_matched_count')" data-col="rules_matched_count">Rules ⇅</th>
          </tr></thead>
          <tbody id="tableBody"></tbody>
        </table>
        <div class="empty" id="emptyState" style="display:none">
          <div class="empty-icon">🔍</div><div>No records match your filters</div>
        </div>
        <div id="loadMoreWrap" style="text-align:center;padding:14px;display:none">
          <button class="btn btn-ghost" onclick="loadMore()">Load More</button>
        </div>
      </div>
    </div>

    <!-- Summary view -->
    <div class="view" id="view-summary">
      <div style="overflow-y:auto;flex:1"><div class="sum-grid" id="summaryGrid"></div></div>
    </div>
  </div>
</div>
</div>

<!-- Detail Panel -->
<div class="detail-panel hidden" id="detailPanel">
  <div class="dp-resize" id="dpResize"></div>
  <div class="dp-header">
    <span class="dp-drag" id="dpDrag" title="Drag">⠿</span>
    <span class="dp-title">Record Detail</span>
    <button class="dp-close" onclick="closeDetail()">✕</button>
  </div>
  <div class="dp-body" id="dpBody">
    <div class="empty"><div class="empty-icon">👆</div><div>Select a row to view details</div></div>
  </div>
</div>
{% endblock %}
{% block scripts %}
<script>
const RUN_ID = {{ run.id }};
const RUN_STATUS = "{{ run.status }}";
let allRows=[], filteredRows=[], currentSev='', sortCol='created', sortDir=-1;
let selectedId=null, detailVisible=false, offset=0, totalRows=0;
const LIMIT=200;
const SEV_COLOR={Critical:'#e74c3c',High:'#e67e22',Medium:'#f1c40f',Low:'#27ae60',Unknown:'#95a5a6'};
const SEV_ORD={Critical:4,High:3,Medium:2,Low:1,'':0};

window.addEventListener('DOMContentLoaded',()=>{
  if(RUN_STATUS==='running'||RUN_STATUS==='pending') pollStatus();
  else { loadResults(); loadSummary(); }
  initDrag(); initResize();
});

// ── Status polling ────────────────────────────────────────────────────────
function pollStatus(){
  document.getElementById('runningBanner').style.display='flex';
  const t=setInterval(async()=>{
    const r=await fetch('/api/run/'+RUN_ID+'/status').then(r=>r.json());
    document.getElementById('progressText').textContent=r.progress||'';
    if(r.db_status==='completed'||r.db_status==='failed'){
      clearInterval(t);
      document.getElementById('runningBanner').style.display='none';
      loadResults(); loadSummary();
    }
  },1500);
}

// ── Filters ───────────────────────────────────────────────────────────────
let searchTimer=null;
function debounceSearch(){ clearTimeout(searchTimer); searchTimer=setTimeout(()=>{ offset=0; loadResults(); },350); }
function filterSev(sev,btn){
  currentSev=sev;
  document.querySelectorAll('.nav-btn').forEach(b=>b.style.borderLeft='');
  if(btn) btn.style.borderLeft='3px solid var(--accent)';
  document.querySelectorAll('.stat-card').forEach(c=>c.classList.remove('active-card'));
  const card=document.querySelector(`.stat-card[data-sev="${sev}"]`);
  if(card) card.classList.add('active-card');
  offset=0; loadResults();
}
function clearFilters(){
  document.getElementById('searchBox').value='';
  document.getElementById('catFilter').value='';
  document.getElementById('nodeFilter').value='';
  currentSev=''; offset=0;
  document.querySelectorAll('.stat-card').forEach(c=>c.classList.remove('active-card'));
  document.querySelector('.stat-card.all').classList.add('active-card');
  loadResults();
}

// ── Load results from API ─────────────────────────────────────────────────
async function loadResults(append=false){
  const search=document.getElementById('searchBox').value;
  const cat=document.getElementById('catFilter').value;
  const node=document.getElementById('nodeFilter').value;
  const params=new URLSearchParams({search,severity:currentSev,category:cat,node,limit:LIMIT,offset});
  const data=await fetch(`/api/run/${RUN_ID}/results?${params}`).then(r=>r.json());
  totalRows=data.total;
  document.getElementById('countBadge').textContent=`${totalRows} record${totalRows!==1?'s':''}`;
  if(append) allRows=[...allRows,...data.results];
  else allRows=[...data.results];
  renderTable();
  document.getElementById('loadMoreWrap').style.display=
    (offset+LIMIT)<totalRows?'block':'none';
}
function loadMore(){ offset+=LIMIT; loadResults(true); }

// ── Populate filter dropdowns ─────────────────────────────────────────────
async function loadSummary(){
  const data=await fetch(`/api/run/${RUN_ID}/summary`).then(r=>r.json());
  // Populate node filter
  const nf=document.getElementById('nodeFilter');
  Object.keys(data.node||{}).sort().forEach(n=>{
    const o=document.createElement('option'); o.value=n; o.textContent=n; nf.appendChild(o);
  });
  // Populate category filter
  const cf=document.getElementById('catFilter');
  Object.keys(data.category||{}).sort().forEach(c=>{
    const o=document.createElement('option'); o.value=c; o.textContent=c; cf.appendChild(o);
  });
  renderSummary(data);
}

// ── Sort ──────────────────────────────────────────────────────────────────
function sortTable(col){
  if(sortCol===col) sortDir*=-1; else { sortCol=col; sortDir=-1; }
  document.querySelectorAll('.data-table thead th').forEach(th=>th.classList.remove('sorted'));
  const th=document.querySelector(`.data-table thead th[data-col="${col}"]`);
  if(th) th.classList.add('sorted');
  allRows.sort((a,b)=>{
    let av=a[col]??'', bv=b[col]??'';
    if(col==='highest_severity'){ av=SEV_ORD[av]??0; bv=SEV_ORD[bv]??0; }
    if(typeof av==='number') return (av-bv)*sortDir;
    return String(av).localeCompare(String(bv))*sortDir;
  });
  renderTable();
}

// ── Render table ──────────────────────────────────────────────────────────
function renderTable(){
  const tbody=document.getElementById('tableBody');
  const empty=document.getElementById('emptyState');
  if(!allRows.length){ tbody.innerHTML=''; empty.style.display='block'; return; }
  empty.style.display='none';
  tbody.innerHTML=allRows.map(r=>{
    const sel=r.id===selectedId?'selected':'';
    const sev=r.highest_severity||'Unknown';
    const msg=esc(r.message||'').substring(0,110)+((r.message||'').length>110?'…':'');
    return `<tr class="${sel}" onclick="selectRow(${r.id})">
      <td style="white-space:nowrap;color:var(--muted);font-size:0.78em">${esc(r.created)}</td>
      <td><span class="sev ${sev}">${sev}</span></td>
      <td style="font-size:0.78em">${esc(r.node)}</td>
      <td style="font-size:0.78em">${esc(r.job_name||'—')}</td>
      <td style="font-size:0.78em">${esc(r.user_name||'—')}</td>
      <td style="font-size:0.76em">${esc(r.categories||'—')}</td>
      <td style="font-size:0.78em;max-width:300px">${msg}</td>
      <td style="text-align:center">
        <span style="background:var(--surface3);border-radius:12px;padding:2px 8px;font-size:0.78em">
          ${r.rules_matched_count}
        </span>
      </td>
    </tr>`;
  }).join('');
}

// ── Select row ────────────────────────────────────────────────────────────
function selectRow(id){
  selectedId=id;
  const r=allRows.find(x=>x.id===id);
  if(!r) return;
  renderTable(); showDetail(r);
}

// ── Detail panel ──────────────────────────────────────────────────────────
function showDetail(r){
  detailVisible=true;
  document.getElementById('detailPanel').classList.remove('hidden');
  document.getElementById('dpToggleBtn').style.background='var(--surface3)';
  const sev=r.highest_severity||'Unknown';
  const ruleIds=(r.matched_rule_ids||'').split(', ').filter(Boolean);
  const ruleNames=(r.matched_rule_names||'').split(', ').filter(Boolean);
  const ruleList=ruleIds.length
    ?ruleIds.map((rid,i)=>`<div class="dp-rec">
        <div class="dp-rec-id">${esc(rid)}</div>
        <div class="dp-rec-name">${esc(ruleNames[i]||'')}</div>
      </div>`).join('')
    :'<div style="color:var(--muted);font-size:0.82em">No rules matched</div>';

  document.getElementById('dpBody').innerHTML=`
    <div class="dp-section">
      <div class="dp-section-title">Overview</div>
      <div class="dp-grid">
        <div class="dp-field"><div class="dp-field-lbl">Severity</div>
          <div class="dp-field-val"><span class="sev ${sev}">${sev}</span></div></div>
        <div class="dp-field"><div class="dp-field-lbl">Timestamp</div>
          <div class="dp-field-val">${esc(r.created)}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">Node</div>
          <div class="dp-field-val">${esc(r.node)}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">Level</div>
          <div class="dp-field-val">${esc(r.level)}</div></div>
        <div class="dp-field full"><div class="dp-field-lbl">Root Cause</div>
          <div class="dp-field-val">${esc(r.root_cause)}</div></div>
      </div>
    </div>
    <div class="dp-section">
      <div class="dp-section-title">Error Message</div>
      <div class="dp-msg">${esc(r.message)}</div>
    </div>
    <div class="dp-section">
      <div class="dp-section-title">Context</div>
      <div class="dp-grid">
        <div class="dp-field"><div class="dp-field-lbl">Job</div><div class="dp-field-val">${esc(r.job_name||'—')}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">User</div><div class="dp-field-val">${esc(r.user_name||'—')}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">Thread</div><div class="dp-field-val">${esc(r.thread||'—')}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">Tx ID</div><div class="dp-field-val">${esc(r.transaction_id||'—')}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">Source</div><div class="dp-field-val">${esc(r.source||'—')}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">Page</div><div class="dp-field-val">${esc(r.page_name||'—')}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">Scope</div><div class="dp-field-val">${esc(r.scope||'—')}</div></div>
        <div class="dp-field"><div class="dp-field-lbl">Script Table</div><div class="dp-field-val">${esc(r.script_table||'—')}</div></div>
        <div class="dp-field full"><div class="dp-field-lbl">Script Record</div><div class="dp-field-val">${esc(r.script_record||'—')}</div></div>
        <div class="dp-field full"><div class="dp-field-lbl">Record Sys ID</div><div class="dp-field-val">${esc(r.record_sys_id||'—')}</div></div>
      </div>
    </div>
    <div class="dp-section">
      <div class="dp-section-title">Categories & Components</div>
      <div class="dp-grid">
        <div class="dp-field full"><div class="dp-field-lbl">Categories</div>
          <div class="dp-field-val">${(r.categories||'').split(', ').map(c=>`<span class="tag">${esc(c)}</span>`).join('')}</div></div>
        <div class="dp-field full"><div class="dp-field-lbl">Components</div>
          <div class="dp-field-val">${(r.affected_components||'').split(', ').map(c=>`<span class="tag">${esc(c)}</span>`).join('')}</div></div>
      </div>
    </div>
    <div class="dp-section">
      <div class="dp-section-title">Recommendations</div>
      <div class="dp-field full" style="margin-top:0">
        <div class="dp-field-val" style="color:#a8d8a8;line-height:1.7">${esc(r.recommendations||'None')}</div>
      </div>
    </div>
    <div class="dp-section">
      <div class="dp-section-title">Matched Rules (${r.rules_matched_count})</div>
      ${ruleList}
    </div>`;
}
function closeDetail(){
  detailVisible=false;
  document.getElementById('detailPanel').classList.add('hidden');
  document.getElementById('dpToggleBtn').style.background='';
  selectedId=null; renderTable();
}
function toggleDetail(){
  if(detailVisible) closeDetail();
  else if(selectedId!==null){ const r=allRows.find(x=>x.id===selectedId); if(r) showDetail(r); }
}

// ── Summary ───────────────────────────────────────────────────────────────
function renderSummary(data){
  const grid=document.getElementById('summaryGrid');
  const sections=[
    {title:'By Severity',data:data.severity,colorKey:'severity'},
    {title:'By Category',data:data.category,colorKey:''},
    {title:'By Node',data:data.node,colorKey:''},
    {title:'By Job',data:data.job,colorKey:''},
  ];
  grid.innerHTML=sections.map(s=>{
    const sorted=Object.entries(s.data||{}).sort((a,b)=>b[1]-a[1]).slice(0,10);
    const max=sorted[0]?.[1]||1;
    const bars=sorted.map(([lbl,c])=>{
      const pct=Math.round(c/max*100);
      const color=s.colorKey==='severity'?(SEV_COLOR[lbl]||'#95a5a6'):'var(--accent)';
      return `<div class="sum-bar-row">
        <div class="sum-bar-lbl" title="${esc(lbl)}">${esc(lbl)}</div>
        <div class="sum-bar-track"><div class="sum-bar-fill" style="width:${pct}%;background:${color}"></div></div>
        <div class="sum-bar-cnt">${c}</div>
      </div>`;
    }).join('');
    return `<div class="sum-card"><h3>${s.title}</h3>${bars}</div>`;
  }).join('');
}

// ── View switch ───────────────────────────────────────────────────────────
function switchView(name,btn){
  document.querySelectorAll('.view').forEach(v=>v.classList.remove('active'));
  document.getElementById('view-'+name).classList.add('active');
  document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
  if(btn) btn.classList.add('active');
}
function toggleSidebar(){ document.getElementById('sidebar').classList.toggle('collapsed'); }

// ── Drag & Resize ─────────────────────────────────────────────────────────
function initDrag(){
  const panel=document.getElementById('detailPanel');
  const handle=document.getElementById('dpDrag');
  let dragging=false,startX=0,startRight=0;
  handle.addEventListener('mousedown',e=>{dragging=true;startX=e.clientX;startRight=parseInt(getComputedStyle(panel).right)||0;panel.classList.add('dragging');e.preventDefault();});
  document.addEventListener('mousemove',e=>{if(!dragging)return;panel.style.right=Math.max(-200,Math.min(window.innerWidth-100,startRight+(startX-e.clientX)))+'px';});
  document.addEventListener('mouseup',()=>{dragging=false;panel.classList.remove('dragging');});
}
function initResize(){
  const panel=document.getElementById('detailPanel');
  const handle=document.getElementById('dpResize');
  let resizing=false,startX=0,startW=0;
  handle.addEventListener('mousedown',e=>{resizing=true;startX=e.clientX;startW=panel.offsetWidth;e.preventDefault();});
  document.addEventListener('mousemove',e=>{if(!resizing)return;panel.style.width=Math.max(300,Math.min(900,startW+(startX-e.clientX)))+'px';});
  document.addEventListener('mouseup',()=>{resizing=false;});
}
function esc(s){ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }
</script>
{% endblock %}
""")

    # ── templates/environments.html ───────────────────────────────────────
    create_file(f"{base}/templates/environments.html", """\
{% extends "base.html" %}
{% block title %}Environments — SN Syslog Analyzer{% endblock %}
{% block content %}
<div class="page">
  <div class="flex mb-2">
    <h2 style="font-size:1.1em;font-weight:700;flex:1">🌐 Environments</h2>
    <button class="btn btn-primary" onclick="openModal()">+ Add Environment</button>
  </div>

  <div class="card">
    {% if envs %}
    <div style="overflow-x:auto">
    <table class="tbl">
      <thead><tr>
        <th>Name</th><th>Instance URL</th><th>Table</th>
        <th>Query Window</th><th>Level</th><th>Active</th><th>Actions</th>
      </tr></thead>
      <tbody>
      {% for e in envs %}
      <tr>
        <td><strong>{{ e.name }}</strong></td>
        <td style="font-size:0.8em;max-width:280px">{{ e.instance_url }}</td>
        <td>{{ e.table_name }}</td>
        <td style="font-size:0.78em">{{ e.start_date }} {{ e.start_time }}<br>→ {{ e.end_date }} {{ e.end_time }}</td>
        <td>{{ e.level }}</td>
        <td>
          {% if e.is_active %}
          <span class="status completed">● Active</span>
          {% else %}
          <button class="btn btn-ghost btn-sm" onclick="activate({{ e.id }})">Set Active</button>
          {% endif %}
        </td>
        <td>
          <div class="flex gap-2">
            <button class="btn btn-ghost btn-sm" onclick="editEnv({{ e.id }})">✏ Edit</button>
            <button class="btn btn-danger btn-sm" onclick="deleteEnv({{ e.id }})">🗑</button>
          </div>
        </td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    </div>
    {% else %}
    <div class="empty"><div class="empty-icon">🌐</div><div>No environments configured.</div></div>
    {% endif %}
  </div>
</div>

<!-- Modal -->
<div class="modal-overlay" id="modalOverlay" onclick="closeModal(event)">
  <div class="modal" onclick="event.stopPropagation()">
    <div class="modal-title" id="modalTitle">Add Environment</div>
    <input type="hidden" id="envId"/>
    <div class="form-row">
      <div class="form-group">
        <label class="form-label">Name *</label>
        <input class="form-control" id="fName" placeholder="e.g. UAT" required/>
      </div>
      <div class="form-group">
        <label class="form-label">Table</label>
        <input class="form-control" id="fTable" value="syslog"/>
      </div>
    </div>
    <div class="form-group">
      <label class="form-label">Instance URL *</label>
      <input class="form-control" id="fUrl" placeholder="https://your-instance.service-now.com" required/>
    </div>
    <div class="form-row">
      <div class="form-group">
        <label class="form-label">Username *</label>
        <input class="form-control" id="fUser" placeholder="username" required/>
      </div>
      <div class="form-group">
        <label class="form-label">Password *</label>
        <input class="form-control" id="fPass" type="password" placeholder="password" required/>
      </div>
    </div>
    <div class="form-row-3">
      <div class="form-group">
        <label class="form-label">Page Size</label>
        <input class="form-control" id="fPageSize" type="number" value="100"/>
      </div>
      <div class="form-group">
        <label class="form-label">Timeout (s)</label>
        <input class="form-control" id="fTimeout" type="number" value="30"/>
      </div>
      <div class="form-group">
        <label class="form-label">Level</label>
        <select class="form-control" id="fLevel">
          <option value="2">2 — Error</option>
          <option value="1">1 — Warning</option>
          <option value="0">0 — Info</option>
        </select>
      </div>
    </div>
    <div style="font-size:0.78em;color:var(--muted);margin-bottom:8px;font-weight:600;text-transform:uppercase;letter-spacing:.8px">Default Query Window</div>
    <div class="form-row">
      <div class="form-group">
        <label class="form-label">Start Date</label>
        <input class="form-control" id="fStartDate" type="date"/>
      </div>
      <div class="form-group">
        <label class="form-label">Start Time</label>
        <input class="form-control" id="fStartTime" type="time" step="1" value="00:00:00"/>
      </div>
      <div class="form-group">
        <label class="form-label">End Date</label>
        <input class="form-control" id="fEndDate" type="date"/>
      </div>
      <div class="form-group">
        <label class="form-label">End Time</label>
        <input class="form-control" id="fEndTime" type="time" step="1" value="23:59:59"/>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-ghost" onclick="closeModal()">Cancel</button>
      <button class="btn btn-primary" onclick="saveEnv()">Save</button>
    </div>
  </div>
</div>
{% endblock %}
{% block scripts %}
<script>
function openModal(){ document.getElementById('modalOverlay').classList.add('open'); }
function closeModal(e){ if(!e||e.target===document.getElementById('modalOverlay')) document.getElementById('modalOverlay').classList.remove('open'); }
function clearForm(){
  ['envId','fName','fUrl','fUser','fPass','fStartDate','fStartTime','fEndDate','fEndTime'].forEach(id=>document.getElementById(id).value='');
  document.getElementById('fTable').value='syslog';
  document.getElementById('fPageSize').value='100';
  document.getElementById('fTimeout').value='30';
  document.getElementById('fLevel').value='2';
  document.getElementById('fStartTime').value='00:00:00';
  document.getElementById('fEndTime').value='23:59:59';
  document.getElementById('modalTitle').textContent='Add Environment';
}
async function editEnv(id){
  const e=await fetch('/api/environments/'+id).then(r=>r.json());
  document.getElementById('envId').value=e.id;
  document.getElementById('fName').value=e.name;
  document.getElementById('fUrl').value=e.instance_url;
  document.getElementById('fUser').value=e.username;
  document.getElementById('fPass').value=e.password;
  document.getElementById('fTable').value=e.table_name;
  document.getElementById('fPageSize').value=e.page_size;
  document.getElementById('fTimeout').value=e.timeout;
  document.getElementById('fLevel').value=e.level;
  document.getElementById('fStartDate').value=e.start_date;
  document.getElementById('fStartTime').value=e.start_time;
  document.getElementById('fEndDate').value=e.end_date;
  document.getElementById('fEndTime').value=e.end_time;
  document.getElementById('modalTitle').textContent='Edit Environment';
  openModal();
}
async function saveEnv(){
  const id=document.getElementById('envId').value;
  const data={
    name:document.getElementById('fName').value,
    instance_url:document.getElementById('fUrl').value,
    username:document.getElementById('fUser').value,
    password:document.getElementById('fPass').value,
    table_name:document.getElementById('fTable').value,
    page_size:document.getElementById('fPageSize').value,
    timeout:document.getElementById('fTimeout').value,
    level:document.getElementById('fLevel').value,
    start_date:document.getElementById('fStartDate').value,
    start_time:document.getElementById('fStartTime').value,
    end_date:document.getElementById('fEndDate').value,
    end_time:document.getElementById('fEndTime').value,
  };
  const url=id?'/api/environments/'+id:'/api/environments';
  const method=id?'PUT':'POST';
  await fetch(url,{method,headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});
  location.reload();
}
async function deleteEnv(id){
  if(!confirm('Delete this environment?')) return;
  await fetch('/api/environments/'+id,{method:'DELETE'});
  location.reload();
}
async function activate(id){
  await fetch('/api/environments/'+id+'/activate',{method:'POST'});
  location.reload();
}
document.getElementById('modalOverlay').addEventListener('click',e=>{
  if(e.target===document.getElementById('modalOverlay')) closeModal();
});
</script>
{% endblock %}
""")

    # ── templates/rules.html ──────────────────────────────────────────────
    create_file(f"{base}/templates/rules.html", """\
{% extends "base.html" %}
{% block title %}Rules — SN Syslog Analyzer{% endblock %}
{% block extra_style %}
.rules-toolbar{display:flex;gap:9px;flex-wrap:wrap;align-items:center;margin-bottom:16px}
.rule-row td{vertical-align:middle}
.rule-row.inactive td{opacity:.45}
.badge{display:inline-block;padding:2px 8px;border-radius:20px;font-size:0.72em;font-weight:700}
.badge-cat{background:var(--surface3);color:var(--text2);border:1px solid var(--border)}
{% endblock %}
{% block content %}
<div class="page">
  <div class="flex mb-2">
    <h2 style="font-size:1.1em;font-weight:700;flex:1">📋 Rules Management
      <span id="ruleCount" style="font-size:0.75em;color:var(--muted);font-weight:400;margin-left:8px"></span>
    </h2>
    <button class="btn btn-primary" onclick="openModal()">+ Add Rule</button>
  </div>

  <!-- Toolbar -->
  <div class="rules-toolbar">
    <div class="search-wrap flex-1" style="max-width:360px">
      <span class="si">🔍</span>
      <input class="form-control" type="text" id="searchBox" placeholder="Search name, pattern, description…" oninput="debounce()"/>
    </div>
    <select class="form-control" id="catFilter" onchange="loadRules()" style="width:auto">
      <option value="">All Categories</option>
    </select>
    <select class="form-control" id="sevFilter" onchange="loadRules()" style="width:auto">
      <option value="">All Severities</option>
      <option>Critical</option><option>High</option><option>Medium</option><option>Low</option>
    </select>
    <label class="flex gap-2" style="font-size:0.85em;color:var(--text2)">
      <input type="checkbox" id="activeOnly" onchange="loadRules()"/> Active only
    </label>
    <button class="btn btn-ghost" onclick="clearFilters()">✕ Clear</button>
  </div>

  <div class="card" style="padding:0;overflow:hidden">
    <div style="overflow-x:auto;max-height:calc(100vh - 260px);overflow-y:auto">
    <table class="tbl" id="rulesTable">
      <thead><tr>
        <th>ID</th><th>Name</th><th>Category</th><th>Severity</th>
        <th>Pattern</th><th>Component</th><th>Active</th><th>Actions</th>
      </tr></thead>
      <tbody id="rulesBody"></tbody>
    </table>
    </div>
    <div class="empty" id="emptyState" style="display:none;padding:40px">
      <div class="empty-icon">📋</div><div>No rules found</div>
    </div>
  </div>
</div>

<!-- Rule Modal -->
<div class="modal-overlay" id="modalOverlay">
  <div class="modal" onclick="event.stopPropagation()" style="max-width:700px">
    <div class="modal-title" id="modalTitle">Add Rule</div>
    <input type="hidden" id="ruleDbId"/>
    <div class="form-row">
      <div class="form-group">
        <label class="form-label">Rule ID *</label>
        <input class="form-control" id="fRuleId" placeholder="RULE_106"/>
      </div>
      <div class="form-group">
        <label class="form-label">Severity *</label>
        <select class="form-control" id="fSeverity">
          <option>Critical</option><option>High</option>
          <option selected>Medium</option><option>Low</option>
        </select>
      </div>
    </div>
    <div class="form-group">
      <label class="form-label">Name *</label>
      <input class="form-control" id="fName" placeholder="Rule name"/>
    </div>
    <div class="form-group">
      <label class="form-label">Description</label>
      <input class="form-control" id="fDesc" placeholder="What this rule detects"/>
    </div>
    <div class="form-group">
      <label class="form-label">Regex Pattern *</label>
      <input class="form-control" id="fPattern" placeholder="e.g. Cannot read property.*from null" style="font-family:monospace"/>
    </div>
    <div class="form-row">
      <div class="form-group">
        <label class="form-label">Category</label>
        <input class="form-control" id="fCategory" placeholder="e.g. NullPointerError" list="catList"/>
        <datalist id="catList"></datalist>
      </div>
      <div class="form-group">
        <label class="form-label">Affected Component</label>
        <input class="form-control" id="fComponent" placeholder="e.g. Script Include"/>
      </div>
    </div>
    <div class="form-group">
      <label class="form-label">Source Hint</label>
      <input class="form-control" id="fSourceHint" placeholder="e.g. ResponseHandler script"/>
    </div>
    <div class="form-group">
      <label class="form-label">Recommendation</label>
      <textarea class="form-control" id="fRec" rows="2" placeholder="What to do when this rule matches"></textarea>
    </div>
    <div class="form-group">
      <label class="form-label">Tags (comma-separated)</label>
      <input class="form-control" id="fTags" placeholder="null, data, responsehandler"/>
    </div>
    <div class="form-group flex gap-2">
      <label class="toggle"><input type="checkbox" id="fActive" checked/><span class="toggle-slider"></span></label>
      <span style="font-size:0.85em;color:var(--text2)">Active</span>
    </div>
    <div class="modal-footer">
      <button class="btn btn-ghost" onclick="closeModal()">Cancel</button>
      <button class="btn btn-primary" onclick="saveRule()">Save Rule</button>
    </div>
  </div>
</div>
{% endblock %}
{% block scripts %}
<script>
let categories=[];
let searchTimer=null;
function debounce(){ clearTimeout(searchTimer); searchTimer=setTimeout(loadRules,300); }

window.addEventListener('DOMContentLoaded',loadRules);

async function loadRules(){
  const search=document.getElementById('searchBox').value;
  const cat=document.getElementById('catFilter').value;
  const sev=document.getElementById('sevFilter').value;
  const active=document.getElementById('activeOnly').checked?'1':'';
  const params=new URLSearchParams({search,category:cat,severity:sev,active_only:active});
  const data=await fetch('/api/rules?'+params).then(r=>r.json());
  categories=data.categories||[];
  // Populate category filter
  const cf=document.getElementById('catFilter');
  const cur=cf.value;
  cf.innerHTML='<option value="">All Categories</option>';
  categories.forEach(c=>{ const o=document.createElement('option'); o.value=c; o.textContent=c; cf.appendChild(o); });
  cf.value=cur;
  // Populate datalist
  const dl=document.getElementById('catList');
  dl.innerHTML=categories.map(c=>`<option value="${c}">`).join('');
  // Update count
  document.getElementById('ruleCount').textContent=`(${data.total} rules)`;
  // Render
  const tbody=document.getElementById('rulesBody');
  const empty=document.getElementById('emptyState');
  if(!data.rules.length){ tbody.innerHTML=''; empty.style.display='block'; return; }
  empty.style.display='none';
  tbody.innerHTML=data.rules.map(r=>`
    <tr class="rule-row ${r.is_active?'':'inactive'}">
      <td><code style="font-size:0.78em;color:var(--accent)">${esc(r.rule_id)}</code></td>
      <td style="max-width:200px"><strong>${esc(r.name)}</strong>
        <div style="font-size:0.75em;color:var(--muted);margin-top:2px">${esc(r.description||'')}</div>
      </td>
      <td><span class="badge badge-cat">${esc(r.category)}</span></td>
      <td><span class="sev ${r.severity}">${r.severity}</span></td>
      <td style="max-width:200px"><code style="font-size:0.75em;color:var(--text2)">${esc(r.pattern)}</code></td>
      <td style="font-size:0.8em">${esc(r.affected_component)}</td>
      <td>
        <label class="toggle">
          <input type="checkbox" ${r.is_active?'checked':''} onchange="toggleRule(${r.id})"/>
          <span class="toggle-slider"></span>
        </label>
      </td>
      <td>
        <div class="flex gap-2">
          <button class="btn btn-ghost btn-sm" onclick="editRule(${r.id})">✏</button>
          <button class="btn btn-danger btn-sm" onclick="deleteRule(${r.id})">🗑</button>
        </div>
      </td>
    </tr>`).join('');
}

function clearFilters(){
  document.getElementById('searchBox').value='';
  document.getElementById('catFilter').value='';
  document.getElementById('sevFilter').value='';
  document.getElementById('activeOnly').checked=false;
  loadRules();
}

function openModal(){ document.getElementById('modalOverlay').classList.add('open'); }
function closeModal(){ document.getElementById('modalOverlay').classList.remove('open'); clearForm(); }
document.getElementById('modalOverlay').addEventListener('click',e=>{
  if(e.target===document.getElementById('modalOverlay')) closeModal();
});

function clearForm(){
  ['ruleDbId','fRuleId','fName','fDesc','fPattern','fCategory','fComponent','fSourceHint','fRec','fTags'].forEach(id=>document.getElementById(id).value='');
  document.getElementById('fSeverity').value='Medium';
  document.getElementById('fActive').checked=true;
  document.getElementById('modalTitle').textContent='Add Rule';
}

async function editRule(id){
  const r=await fetch('/api/rules/'+id).then(r=>r.json());
  document.getElementById('ruleDbId').value=r.id;
  document.getElementById('fRuleId').value=r.rule_id;
  document.getElementById('fName').value=r.name;
  document.getElementById('fDesc').value=r.description||'';
  document.getElementById('fPattern').value=r.pattern;
  document.getElementById('fCategory').value=r.category||'';
  document.getElementById('fSeverity').value=r.severity||'Medium';
  document.getElementById('fComponent').value=r.affected_component||'';
  document.getElementById('fSourceHint').value=r.source_hint||'';
  document.getElementById('fRec').value=r.recommendation||'';
  let tags=r.tags||'[]';
  try{ tags=JSON.parse(tags).join(', '); }catch(e){ tags=tags; }
  document.getElementById('fTags').value=tags;
  document.getElementById('fActive').checked=!!r.is_active;
  document.getElementById('modalTitle').textContent='Edit Rule';
  openModal();
}

async function saveRule(){
  const id=document.getElementById('ruleDbId').value;
  const tagsRaw=document.getElementById('fTags').value;
  const tags=tagsRaw?tagsRaw.split(',').map(t=>t.trim()).filter(Boolean):[];
  const data={
    rule_id:document.getElementById('fRuleId').value,
    name:document.getElementById('fName').value,
    description:document.getElementById('fDesc').value,
    pattern:document.getElementById('fPattern').value,
    category:document.getElementById('fCategory').value,
    severity:document.getElementById('fSeverity').value,
    affected_component:document.getElementById('fComponent').value,
    source_hint:document.getElementById('fSourceHint').value,
    recommendation:document.getElementById('fRec').value,
    tags:tags,
    is_active:document.getElementById('fActive').checked,
  };
  if(!data.rule_id||!data.name||!data.pattern){ alert('Rule ID, Name and Pattern are required.'); return; }
  const url=id?'/api/rules/'+id:'/api/rules';
  const method=id?'PUT':'POST';
  await fetch(url,{method,headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});
  closeModal(); loadRules();
}

async function deleteRule(id){
  if(!confirm('Delete this rule?')) return;
  await fetch('/api/rules/'+id,{method:'DELETE'});
  loadRules();
}

async function toggleRule(id){
  await fetch('/api/rules/'+id+'/toggle',{method:'POST'});
  loadRules();
}

function esc(s){ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }
</script>
{% endblock %}
""")

    # ── templates/settings.html ───────────────────────────────────────────
    create_file(f"{base}/templates/settings.html", """\
{% extends "base.html" %}
{% block title %}Settings — SN Syslog Analyzer{% endblock %}
{% block content %}
<div class="page" style="max-width:700px">
  <h2 style="font-size:1.1em;font-weight:700;margin-bottom:16px">⚙ Settings</h2>
  <div class="card">
    <form method="POST" action="/settings">
      <div class="form-group">
        <label class="form-label">Output Folder</label>
        <input class="form-control" name="output_folder" value="{{ settings.get('output_folder','output') }}"/>
      </div>
      <div class="form-group">
        <label class="form-label">Output Filename Prefix</label>
        <input class="form-control" name="output_prefix" value="{{ settings.get('output_prefix','output_error_') }}"/>
      </div>
      <div class="form-group">
        <label class="form-label">Output Formats (comma-separated: xlsx, html)</label>
        <input class="form-control" name="output_formats" value="{{ settings.get('output_formats','xlsx,html') }}"/>
      </div>
      <div class="form-group">
        <label class="form-label">Log Level</label>
        <select class="form-control" name="log_level">
          {% for lvl in ['DEBUG','INFO','WARNING','ERROR'] %}
          <option {% if settings.get('log_level','INFO')==lvl %}selected{% endif %}>{{ lvl }}</option>
          {% endfor %}
        </select>
      </div>
      <div class="modal-footer" style="border:none;padding:0;margin-top:20px">
        <button class="btn btn-primary" type="submit">💾 Save Settings</button>
      </div>
    </form>
  </div>

  <div class="card mt-2">
    <div class="card-title">📦 Database Info</div>
    <div class="text-muted">
      <div>Database: <code>syslog_analyzer.db</code></div>
      <div class="mt-1">All environments, rules, settings and analysis results are stored in SQLite.</div>
    </div>
  </div>
</div>
{% endblock %}
""")

    # ── Done ──────────────────────────────────────────────────────────────
    print("\n" + "=" * 65)
    print("  ✅  Flask + SQLite project generated successfully!")
    print("=" * 65)
    print(f"""
  📁 {base}/
  ├── app.py                 ← Flask application (main entry)
  ├── database.py            ← SQLite ORM + seed data
  ├── analyzer.py            ← Rule-based log analyzer
  ├── sn_client.py           ← ServiceNow Table API client
  ├── reporter.py            ← XLSX report generator
  ├── requirements.txt
  ├── templates/
  │   ├── base.html          ← Shared layout + dark theme
  │   ├── dashboard.html     ← Run history + start run form
  │   ├── run_detail.html    ← Live results viewer
  │   ├── environments.html  ← CRUD environments
  │   ├── rules.html         ← CRUD rules + search
  │   └── settings.html      ← App settings
  └── output/                ← XLSX reports saved here

  ▶  Steps:
     1.  cd {base}
     2.  pip install -r requirements.txt
     3.  python app.py
     4.  Open http://localhost:5000

  🗄  SQLite DB: syslog_analyzer.db (auto-created on first run)
     • 3 default environments (DEV / UAT / PROD)
     • 105 seeded rules
     • All CRUD via the web UI

  🌐  Pages:
     /                  → Dashboard + run history
     /environments      → Manage environments (CRUD)
     /rules             → Manage rules (CRUD + search + toggle)
     /settings          → App settings
     /run/<id>          → Live results with detail panel
""")


if __name__ == "__main__":
    generate_project()
